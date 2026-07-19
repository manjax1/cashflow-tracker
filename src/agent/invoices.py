"""Invoice Intelligence — Session 1: extraction & classification.

Parses invoices dropped into invoices_inbox/ (Amazon order PDFs, .eml, .txt),
uses the LLM to extract structured order data and classify each item against
YOUR existing category taxonomy, then deterministically allocates tax/
promotions so item amounts sum exactly to the grand total.

Design rule (same as the agent): the LLM extracts and proposes; deterministic
code validates, allocates, and stores. Nothing touches the ledger here —
matching & splits are Session 2.

Usage:
    python -m src.agent.invoices ingest --dry-run   # parse+classify, print only
    python -m src.agent.invoices ingest             # also save JSON to invoices_data/
    python -m src.agent.invoices list               # show extracted orders
"""

import json
import os
import re
import sys
from datetime import datetime

from dotenv import load_dotenv

load_dotenv()

import anthropic  # noqa: E402

from . import ledger  # noqa: E402
from .tools import audit  # noqa: E402

INBOX = os.path.join(ledger.REPO_ROOT, "invoices_inbox")
OUTDIR = os.path.join(ledger.REPO_ROOT, "invoices_data")
MODEL = os.environ.get("AGENT_MODEL", "claude-sonnet-4-5")

EXTRACT_TOOL = {
    "name": "record_order",
    "description": "Record the structured contents of one retail invoice/order.",
    "input_schema": {
        "type": "object",
        "properties": {
            "order_id": {"type": "string"},
            "order_date": {"type": "string", "description": "ISO date YYYY-MM-DD"},
            "merchant": {"type": "string", "description": "e.g. Amazon"},
            "items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string"},
                        "price": {"type": "number", "description": "unit price"},
                        "quantity": {"type": "integer", "default": 1},
                        "category": {"type": "string",
                                     "description": "MUST be one of the provided taxonomy categories"},
                        "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
                    },
                    "required": ["name", "price", "category", "confidence"],
                },
            },
            "subtotal": {"type": "number"},
            "shipping": {"type": "number", "default": 0},
            "promotion": {"type": "number", "description": "discount as NEGATIVE number, 0 if none"},
            "tax": {"type": "number", "default": 0},
            "grand_total": {"type": "number", "description": "the amount actually charged"},
            "payment_hint": {"type": "string", "description": "e.g. 'Visa ending 3810'"},
            "notes": {"type": "string", "description": "anything odd: multiple shipments, gift cards, ambiguity"},
        },
        "required": ["order_id", "order_date", "merchant", "items", "grand_total"],
    },
}

SYSTEM = """You extract structured data from retail invoices and classify items.
Classification rules:
- Use ONLY the category names provided in the taxonomy list. Never invent one.
- Food/household consumables -> Groceries; plants/garden supplies -> Gardening;
  apparel/shoes -> Clothing; skincare, hygiene, grooming, cosmetics, vitamins &
  supplements -> Personal Care; gym/sports/therapy equipment -> Health and
  Fitness; devices/accessories/electronics -> the best-fitting existing
  category; if truly unclear use 'Other - Uncategorized' with low confidence.
- Mark confidence honestly: 'high' only when the item name is unambiguous.
- Report money fields exactly as printed. Do not compute or adjust totals.
- Watch for quantities: a lone number near an item line usually means quantity;
  verify item prices x quantities roughly reconcile with the printed subtotal.
- If an item shows a return/refund in progress, keep it in the items list (it
  was part of the charge) and mention the pending refund in notes.
- If the order was paid with rewards points or gift balance (Grand Total $0 or
  a large points deduction), say so in notes."""


# ----------------------------- file parsing -----------------------------

def read_pdf(path):
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def read_eml(path):
    import email
    from email import policy
    with open(path, "rb") as f:
        msg = email.message_from_binary_file(f, policy=policy.default)
    body = msg.get_body(preferencelist=("html", "plain"))
    text = body.get_content() if body else ""
    text = re.sub(r"<(script|style).*?</\1>", " ", text, flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", " ", text)          # strip tags
    return re.sub(r"\s+", " ", text)


def read_invoice(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return read_pdf(path)
    if ext == ".eml":
        return read_eml(path)
    with open(path, errors="replace") as f:
        return f.read()


# --------------------------- LLM extraction ---------------------------

# Categories that should exist for item-level classification even before any
# ledger transaction uses them (they self-register once splits land).
EXTRA_CATEGORIES = ["Personal Care"]


def taxonomy():
    cats = {c["category"] for c in ledger.list_categories()["categories"]}
    cats.update(EXTRA_CATEGORIES)
    return sorted(cats)


def extract_order(text, client=None):
    client = client or anthropic.Anthropic()
    cats = taxonomy()
    prompt = (f"Taxonomy categories (use these EXACT names):\n"
              + "\n".join(f"- {c}" for c in cats)
              + f"\n\nInvoice text:\n---\n{text[:12000]}\n---\n"
              "Extract this order with record_order.")
    resp = client.messages.create(
        model=MODEL, max_tokens=4000, system=SYSTEM,
        tools=[EXTRACT_TOOL], tool_choice={"type": "tool", "name": "record_order"},
        messages=[{"role": "user", "content": prompt}])
    order = next(b.input for b in resp.content if b.type == "tool_use")
    order["_tokens"] = {"in": resp.usage.input_tokens, "out": resp.usage.output_tokens}
    return order


# ----------------------- deterministic allocation -----------------------

def allocate(order):
    """Scale item lines so allocated amounts sum EXACTLY to grand_total
    (tax, shipping, promotions spread proportionally; remainder pennies to
    the largest item). Flags a validation warning if inputs look off."""
    items = order.get("items", [])
    gross = [round(i["price"] * i.get("quantity", 1), 2) for i in items]
    base = sum(gross)
    total = order["grand_total"]
    warnings = []
    if not items or base <= 0:
        order["allocation_warning"] = "no items or zero subtotal"
        return order
    if total == 0:
        # Paid entirely with rewards points / gift card: no card charge exists,
        # so there is nothing to match against the ledger.
        for item, g in zip(items, gross):
            item["gross"] = g
            item["allocated_amount"] = 0.0
        order["no_card_charge"] = True
        order["notes"] = ((order.get("notes") or "") +
                          " Paid with rewards/gift balance — no ledger charge to match.").strip()
        return order
    printed = order.get("subtotal")
    if printed is not None and abs(printed - base) > 0.02:
        warnings.append(f"item sum {base:.2f} != printed subtotal {printed:.2f}")
    alloc = [round(g * total / base, 2) for g in gross]
    drift = round(total - sum(alloc), 2)
    if drift:
        alloc[gross.index(max(gross))] = round(alloc[gross.index(max(gross))] + drift, 2)
    for item, g, a in zip(items, gross, alloc):
        item["gross"] = g
        item["allocated_amount"] = a
    if warnings:
        order["allocation_warning"] = "; ".join(warnings)
    return order


# ------------------------------ pipeline ------------------------------

def order_path(order_id):
    return os.path.join(OUTDIR, f"{order_id}.json")


def ingest(dry_run=False):
    if not os.path.isdir(INBOX):
        sys.exit(f"No inbox folder at {INBOX}")
    files = sorted(f for f in os.listdir(INBOX)
                   if os.path.splitext(f)[1].lower() in (".pdf", ".eml", ".txt", ".html"))
    if not files:
        sys.exit(f"No invoice files (.pdf/.eml/.txt/.html) in {INBOX}")
    os.makedirs(OUTDIR, exist_ok=True)
    client = anthropic.Anthropic()
    done = skipped = 0
    for fname in files:
        path = os.path.join(INBOX, fname)
        text = read_invoice(path)
        m = re.search(r"Order\s*#?\s*([\d-]{15,25})", text) or re.search(r"([\d]{3}-[\d]{7}-[\d]{7})", fname)
        oid_hint = m.group(1) if m else None
        if oid_hint and os.path.exists(order_path(oid_hint)):
            print(f"  ↷ {fname}: already extracted ({oid_hint}), skipping")
            skipped += 1
            continue
        print(f"  ⚙ extracting {fname} ...")
        order = allocate(extract_order(text, client))
        order["_source_file"] = fname
        order["_extracted_at"] = datetime.now().isoformat(timespec="seconds")
        print_order(order)
        if not dry_run:
            with open(order_path(order["order_id"]), "w") as f:
                json.dump(order, f, indent=2)
            audit("invoice_extracted", {"order_id": order["order_id"],
                                        "file": fname, "total": order["grand_total"]})
        done += 1
    print(f"\n{done} extracted, {skipped} skipped"
          + (" (dry run — nothing saved)" if dry_run else f" → {OUTDIR}"))


def print_order(o):
    tag = "  ★ NO CARD CHARGE (rewards-paid)" if o.get("no_card_charge") else ""
    print(f"    {o['merchant']} order {o['order_id']}  ({o['order_date']})  "
          f"total ${o['grand_total']:.2f}  [{o.get('payment_hint', '?')}]{tag}")
    for i in o["items"]:
        conf = {"high": "", "medium": " (?)", "low": " (??)"}[i["confidence"]]
        print(f"      ${i.get('allocated_amount', i['price']):>8.2f}  "
              f"{i['category']}{conf}  —  {i['name'][:58]}")
    if o.get("allocation_warning"):
        print(f"      ⚠ {o['allocation_warning']}")
    if o.get("notes"):
        print(f"      note: {o['notes']}")


# ----------------------- Amazon data-export adapter -----------------------

EXPORT_CSV = os.path.join(OUTDIR, "amazon_export", "Your Amazon Orders", "Order History.csv")
CACHE_PATH = os.path.join(OUTDIR, "classification_cache.json")

CLASSIFY_TOOL = {
    "name": "classify_items",
    "description": "Classify retail product names into the given taxonomy.",
    "input_schema": {
        "type": "object",
        "properties": {
            "classifications": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "product name EXACTLY as given"},
                        "category": {"type": "string"},
                        "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
                    },
                    "required": ["name", "category", "confidence"],
                },
            }
        },
        "required": ["classifications"],
    },
}


def money(v):
    try:
        return float(str(v).strip().strip("'\""))
    except (ValueError, TypeError):
        return 0.0


def classify_names(names, client):
    """Batch-classify unique product names with a persistent cache."""
    cache = {}
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH) as f:
            cache = json.load(f)
    todo = [n for n in names if n not in cache]
    cats = taxonomy()
    for i in range(0, len(todo), 40):
        batch = todo[i:i + 40]
        print(f"  ⚙ classifying {len(batch)} product names "
              f"({i + len(batch)}/{len(todo)}) ...")
        prompt = ("Taxonomy categories (use these EXACT names):\n"
                  + "\n".join(f"- {c}" for c in cats)
                  + "\n\nClassify each product:\n"
                  + "\n".join(f"- {n}" for n in batch))
        resp = client.messages.create(
            model=MODEL, max_tokens=4000, system=SYSTEM,
            tools=[CLASSIFY_TOOL],
            tool_choice={"type": "tool", "name": "classify_items"},
            messages=[{"role": "user", "content": prompt}])
        out = next(b.input for b in resp.content if b.type == "tool_use")
        for c in out["classifications"]:
            cache[c["name"]] = {"category": c["category"], "confidence": c["confidence"]}
    with open(CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)
    return cache


def import_export(dry_run=False):
    """Convert Amazon's data-export Order History.csv into the same per-order
    JSON the invoice extractor produces. CSV money is authoritative (discounts
    and tax already baked into per-item Total Amount) — the LLM only
    classifies product names."""
    import csv as csvmod
    from collections import defaultdict

    if not os.path.exists(EXPORT_CSV):
        sys.exit(f"Export CSV not found at {EXPORT_CSV}")
    ledger_start = min(t["Date"] for t in ledger.load_transactions())
    with open(EXPORT_CSV, newline="", encoding="utf-8-sig") as f:
        rows = [r for r in csvmod.DictReader(f)
                if r["Order Date"][:10] >= ledger_start
                and "cancel" not in r["Order Status"].lower()]
    orders = defaultdict(list)
    for r in rows:
        orders[r["Order ID"]].append(r)
    print(f"{len(rows)} item rows in {len(orders)} orders since {ledger_start}")

    names = sorted({r["Product Name"] for r in rows})
    client = anthropic.Anthropic()
    cache = classify_names(names, client)

    os.makedirs(OUTDIR, exist_ok=True)
    saved = skipped = no_charge = 0
    for oid, lines in sorted(orders.items(), key=lambda kv: kv[1][0]["Order Date"]):
        if os.path.exists(order_path(oid)):
            skipped += 1
            continue
        items, shipments = [], {}
        for r in lines:
            cls = cache.get(r["Product Name"], {"category": "Other - Uncategorized",
                                                "confidence": "low"})
            items.append({
                "name": r["Product Name"],
                "price": money(r["Unit Price"]),
                "quantity": int(r["Original Quantity"] or 1),
                "category": cls["category"],
                "confidence": cls["confidence"],
                "gross": money(r["Unit Price"]) * int(r["Original Quantity"] or 1),
                "allocated_amount": money(r["Total Amount"]),  # authoritative
            })
            track = r.get("Carrier Name & Tracking Number", "") or f"ship-{len(shipments)}"
            shipments[track] = round(
                money(r["Shipment Item Subtotal"]) + money(r["Shipment Item Subtotal Tax"]), 2)
        total = round(sum(i["allocated_amount"] for i in items), 2)
        order = {
            "order_id": oid,
            "order_date": lines[0]["Order Date"][:10],
            "merchant": "Amazon",
            "items": items,
            "grand_total": total,
            "payment_hint": lines[0]["Payment Method Type"],
            "shipments": [{"amount": a} for a in shipments.values()],
            "_source_file": "amazon-data-export",
            "_extracted_at": datetime.now().isoformat(timespec="seconds"),
        }
        if total == 0:
            order["no_card_charge"] = True
            order["notes"] = "Paid with rewards/gift balance — no ledger charge to match."
            no_charge += 1
        if "gift" in order["payment_hint"].lower() and total > 0:
            order["notes"] = ("Partially paid with gift balance — card charge may be "
                              "less than order total; matcher should treat amount as approximate.")
        if dry_run and saved < 5:
            print_order(order)
        if not dry_run:
            with open(order_path(oid), "w") as f:
                json.dump(order, f, indent=2)
        saved += 1
    audit("export_imported", {"orders": saved, "skipped": skipped})
    print(f"\n{saved} orders {'previewed' if dry_run else 'saved'}, "
          f"{skipped} already present, {no_charge} rewards-paid (no charge)"
          + (" — dry run, nothing written" if dry_run else f" → {OUTDIR}"))


# ------------------- apply: import orders as ledger rows -------------------

REFUNDS_CSV = os.path.join(OUTDIR, "amazon_export", "Your Returns & Refunds",
                           "Refund Details.csv")
AMAZON_ACCOUNT = "Amazon Visa"


def load_card_orders():
    orders = []
    for fn in os.listdir(OUTDIR):
        if fn.endswith(".json") and fn != os.path.basename(CACHE_PATH):
            with open(os.path.join(OUTDIR, fn)) as f:
                o = json.load(f)
            if isinstance(o, dict) and "order_id" in o and not o.get("no_card_charge"):
                orders.append(o)
    return sorted(orders, key=lambda o: o["order_date"])


def build_rows(orders, existing_refs, ledger_start):
    """Item rows + refund rows for the virtual Amazon Visa account."""
    rows = []
    for o in orders:
        for i, item in enumerate(o["items"], 1):
            ref = f"amazon-{o['order_id']}-{i}"
            amt = round(item.get("allocated_amount", 0), 2)
            if ref in existing_refs or amt == 0:
                continue
            rows.append([o["order_date"], f"Amazon: {item['name'][:80]}",
                         AMAZON_ACCOUNT, item["category"], "Expense", amt,
                         True, ref])
    if os.path.exists(REFUNDS_CSV):
        import csv as csvmod
        oids = {o["order_id"] for o in orders}
        with open(REFUNDS_CSV, newline="", encoding="utf-8-sig") as f:
            for n, r in enumerate(csvmod.DictReader(f), 1):
                date = (r.get("Refund Date") or r.get("Creation Date") or "")[:10]
                amt = money(r.get("Refund Amount"))
                ref = f"amazon-refund-{r['Order ID']}-{n}"
                if (r["Order ID"] in oids and date >= ledger_start and amt > 0
                        and ref not in existing_refs
                        and "reversed" not in (r.get("Reversal Status") or "").lower()):
                    rows.append([date, f"Amazon refund: order {r['Order ID']}",
                                 AMAZON_ACCOUNT, "Income - Retail Refunds",
                                 "Income", amt, True, ref])
    return rows


def apply_orders(dry_run=False):
    """Import Amazon order items as spend rows in the 'Amazon Visa' account,
    and reclassify Chase autopay lumps to the excluded 'Credit Card Payment'
    category (same dedup convention as the BofA card) to avoid double-counting."""
    import shutil
    import openpyxl
    from collections import defaultdict

    txns = ledger.load_transactions()
    ledger_start = min(t["Date"] for t in txns)
    existing_refs = {str(t["SourceRef"]) for t in txns}
    orders = load_card_orders()
    rows = build_rows(orders, existing_refs, ledger_start)
    autopay = [t for t in txns if "CHASE CREDIT CRD DES:AUTOPAY" in t["Description"]
               and t["Category"] != "Credit Card Payment"]

    by_cat = defaultdict(float)
    for r in rows:
        if r[4] == "Expense":
            by_cat[r[3]] += r[5]
    exp_total = sum(v for v in by_cat.values())
    ref_total = sum(r[5] for r in rows if r[4] == "Income")
    print(f"{len(orders)} orders -> {len(rows)} new ledger rows "
          f"(spend ${exp_total:.2f}, refunds ${ref_total:.2f})")
    print(f"{len(autopay)} Chase autopay rows to flip to 'Credit Card Payment' "
          f"(${sum(t['Amount'] for t in autopay):.2f} removed from spend)")
    print("\nNew spend by category:")
    for c, v in sorted(by_cat.items(), key=lambda kv: -kv[1]):
        print(f"  {c:<28} ${v:>9.2f}")
    if dry_run:
        print("\nSample rows:")
        for r in rows[:6]:
            print("  ", r[:6])
        print("\n(dry run — nothing written)")
        return

    confirm = input(f"\nAppend {len(rows)} rows and flip {len(autopay)} autopay rows? [y/N] ")
    if confirm.strip().lower() != "y":
        print("Cancelled.")
        return
    backup = ledger.LEDGER_PATH.replace(
        ".xlsx", f"_BACKUP_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx")
    shutil.copy2(ledger.LEDGER_PATH, backup)
    wb = openpyxl.load_workbook(ledger.LEDGER_PATH)
    ws = wb["Transactions"]
    header = [c.value for c in ws[1]]
    ref_col = header.index("SourceRef")
    cat_col = header.index("Category")
    net_col = header.index("IncludeInNet")
    flip_refs = {str(t["SourceRef"]) for t in autopay}
    flipped = 0
    for row in ws.iter_rows(min_row=2):
        if str(row[ref_col].value) in flip_refs:
            row[cat_col].value = "Credit Card Payment"
            row[net_col].value = False
            flipped += 1
    for r in rows:
        ws.append(r)
    wb.save(ledger.LEDGER_PATH)
    ledger._cache["mtime"] = None
    audit("amazon_orders_applied",
          {"rows_added": len(rows), "autopay_flipped": flipped,
           "spend": round(exp_total, 2), "refunds": round(ref_total, 2),
           "backup": os.path.basename(backup)})
    print(f"✅ {len(rows)} rows appended, {flipped} autopay rows flipped. "
          f"Backup: {os.path.basename(backup)}")
    print("Remember: 'push' from the agent CLI to sync Drive, and re-run "
          "scripts/compact_rules.py for Railway (rules changed).")


PERSONAL_CARE_HINTS = [
    "razor", "shav", "moistur", "lotion", "cream", "serum", "soap", "shampoo",
    "conditioner", "toothpaste", "toothbrush", "floss", "deodorant", "sunscreen",
    "cosmetic", "makeup", "nail", "eyebrow", "scissors", "trimmer", "hygiene",
    "vitamin", "supplement", "glucosamine", "wipes", "facial", "skin", "hair",
]


def reclassify(clear_hints=None):
    """Clear cached classifications matching hint keywords, reclassify those
    names with the (updated) taxonomy, and rewrite saved order JSONs."""
    hints = clear_hints or PERSONAL_CARE_HINTS
    cache = {}
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH) as f:
            cache = json.load(f)
    cleared = [n for n in cache if any(h in n.lower() for h in hints)]
    for n in cleared:
        del cache[n]
    with open(CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)
    print(f"cleared {len(cleared)} cached classifications for reclassification")

    # collect all item names across saved orders; classify anything uncached
    orders = {}
    for fn in os.listdir(OUTDIR):
        if fn.endswith(".json") and fn != os.path.basename(CACHE_PATH):
            with open(os.path.join(OUTDIR, fn)) as fh:
                data = json.load(fh)
            if isinstance(data, dict) and "order_id" in data:
                orders[fn] = data
    names = sorted({i["name"] for o in orders.values() for i in o["items"]})
    cache = classify_names(names, anthropic.Anthropic())

    changed = 0
    for fn, o in orders.items():
        dirty = False
        for i in o["items"]:
            c = cache.get(i["name"])
            if c and (i["category"] != c["category"] or i["confidence"] != c["confidence"]):
                print(f"  {i['category']:<24} -> {c['category']:<18} {i['name'][:48]}")
                i["category"], i["confidence"] = c["category"], c["confidence"]
                dirty = True
                changed += 1
        if dirty:
            with open(os.path.join(OUTDIR, fn), "w") as f:
                json.dump(o, f, indent=2)
    audit("reclassified", {"cleared": len(cleared), "changed": changed})
    print(f"\n{changed} item classifications updated across saved orders")


def list_orders():
    if not os.path.isdir(OUTDIR):
        sys.exit("Nothing extracted yet.")
    orders = []
    for f in sorted(os.listdir(OUTDIR)):
        if not f.endswith(".json"):
            continue
        with open(os.path.join(OUTDIR, f)) as fh:
            data = json.load(fh)
        if isinstance(data, dict) and "order_id" in data:  # skip cache/other json
            orders.append(data)
    for o in sorted(orders, key=lambda x: x["order_date"]):
        print_order(o)
    print(f"\n{len(orders)} orders")


if __name__ == "__main__":
    args = sys.argv[1:]
    if args and args[0] == "ingest":
        ingest(dry_run="--dry-run" in args)
    elif args and args[0] == "import-export":
        import_export(dry_run="--dry-run" in args)
    elif args and args[0] == "reclassify":
        reclassify()
    elif args and args[0] == "apply":
        apply_orders(dry_run="--dry-run" in args)
    elif args and args[0] == "list":
        list_orders()
    else:
        print(__doc__)
