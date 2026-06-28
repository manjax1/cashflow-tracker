#!/usr/bin/env python3
"""
Re-apply current categorization rules to every Transactions row, updating
only the Category column where the result differs from what's stored.

DRY-RUN by default — nothing is written until --apply is passed.

Manual-override protection (requirement 6)
------------------------------------------
Two protection mechanisms exist in the codebase:

1. IncludeInNet = False  (column G)
   Written by write_spending_ledger() when categorize() returns
   exclude_from_net=True, which fires only for the category
   "One-Off - Non-Recurring (excluded from Net)".  These rows were
   deliberately hand-classified and must never be auto-recategorized.

2. PROTECTED_CATEGORIES constant below
   Belt-and-suspenders: also skips any row whose current Category
   string is in this set, even if IncludeInNet is somehow True.

All other categories in the Transactions sheet were set by the pipeline
automatically and are fair game for update.  "Credit Card Payment" and
"Internal Transfer" rows have excluded=True and are never written to the
Transactions sheet at all, so they can't appear here.

PFC data-loss guard
--------------------
Plaid's personal_finance_category (PFC) is not stored in the ledger.
Re-running categorize() with empty PFC yields "Other - Uncategorized"
for any row originally categorized via Plaid's PFC signal (e.g. Dining
from FOOD_AND_DRINK).  This script suppresses any apparent "change"
where the new result is "Other - Uncategorized" but the current stored
category is something more specific — that would silently downgrade
good data, not fix anything.

Usage
-----
    python src/recategorize_ledger.py           # dry-run: shows what would change
    python src/recategorize_ledger.py --apply   # writes changes, verifies after
"""

import argparse
import os
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from filters import load_rules, categorize
from utils import resolve_ledger_path

RULES_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "spending_rules.json")

W = 112   # report line width

# Rows whose current Category is in this set are never touched.
# IncludeInNet=False rows are also always skipped (see module docstring).
PROTECTED_CATEGORIES = frozenset({
    "One-Off - Non-Recurring (excluded from Net)",
})


# ── helpers ──────────────────────────────────────────────────────────────────

def _date_str(val) -> str:
    if hasattr(val, "date"):
        return val.date().isoformat()
    return str(val)[:10]


def _plaid_amount(ledger_amount: float, tx_type: str) -> float:
    """Reconstruct Plaid-convention signed amount from ledger storage.

    Ledger stores abs(amount) in column F; Type column says Income/Expense.
    Plaid convention: positive = expense/debit, negative = income/credit.
    """
    return -abs(ledger_amount) if tx_type == "Income" else abs(ledger_amount)


# ── scan pass (always read-only) ──────────────────────────────────────────────

def _scan(ledger_path: str, rules: list) -> tuple[list, dict]:
    """Open ledger read-only; return (changes, stats).

    changes: list of dicts — one per row that would be updated.
    stats:   counters and before-totals.
    """
    wb = openpyxl.load_workbook(ledger_path, read_only=True, data_only=True)
    ws = wb["Transactions"]

    # Map column names → 0-based indices from the real header row.
    raw_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(raw_header) if name}

    DATE = col["Date"]
    DESC = col["Description"]
    ACCT = col["Account"]
    CAT  = col["Category"]
    TYPE = col["Type"]
    AMT  = col["Amount"]
    NET  = col["IncludeInNet"]

    changes: list[dict] = []
    stats = {
        "rows_total":           0,
        "skipped_protected":    0,
        "skipped_new_excluded": 0,
        "skipped_pfc_guard":    0,
        "income_before":        0.0,
        "expense_before":       0.0,
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[DATE]:
            continue

        current_cat    = str(row[CAT]  or "")
        tx_type        = str(row[TYPE] or "")
        include_in_net = row[NET]
        amt_raw        = float(row[AMT] or 0)

        stats["rows_total"] += 1

        # Accumulate before-totals (mirrors /ledger/info logic exactly)
        if include_in_net is not False:
            if tx_type == "Income":
                stats["income_before"] += amt_raw
            elif tx_type == "Expense":
                stats["expense_before"] += amt_raw

        # --- protection checks ---

        # 1. IncludeInNet=False  →  One-Off / manual override, never touch
        if include_in_net is False:
            stats["skipped_protected"] += 1
            continue

        # 2. Category itself is in the protected set (belt-and-suspenders)
        if current_cat in PROTECTED_CATEGORIES:
            stats["skipped_protected"] += 1
            continue

        # --- reconstruct transaction dict ---

        date_str = _date_str(row[DATE])
        desc     = str(row[DESC] or "")
        acct     = str(row[ACCT] or "")
        amount   = _plaid_amount(amt_raw, tx_type)

        tx = {
            "name":                      desc,
            "amount":                    amount,
            "date":                      date_str,
            "pending":                   False,
            "personal_finance_category": {"primary": ""},  # not stored in ledger
        }

        result = categorize(tx, rules, account_label=acct)

        # --- result filters ---

        # If the new result would exclude the row entirely (e.g. a new Internal
        # Transfer rule now matches), we can't handle that by changing Category
        # alone — IncludeInNet would also need updating.  Flag for manual review.
        if result.get("excluded"):
            stats["skipped_new_excluded"] += 1
            continue

        new_cat = result.get("category", "")

        # PFC data-loss guard: if the only apparent "change" is PFC → empty,
        # the new result falls to "Other - Uncategorized" but the current category
        # was legitimately set by Plaid PFC (e.g. Dining from FOOD_AND_DRINK).
        # Suppressing this prevents silently downgrading good data.
        if new_cat == "Other - Uncategorized" and current_cat not in (
            "Other - Uncategorized", ""
        ):
            stats["skipped_pfc_guard"] += 1
            continue

        if new_cat != current_cat:
            changes.append({
                "row":     row_idx,
                "date":    date_str,
                "desc":    desc[:65],
                "amount":  amt_raw,
                "type":    tx_type,
                "old_cat": current_cat,
                "new_cat": new_cat,
            })

    wb.close()
    return changes, stats


# ── apply pass ────────────────────────────────────────────────────────────────

def _apply(ledger_path: str, changes: list, cat_col_1idx: int) -> None:
    """Open ledger in write mode, update only Category cells, save.

    Uses plain load_workbook() (no read_only, no data_only) — same pattern
    as write_spending_ledger() — so formulas in Monthly Summary / YTD /
    YoY sheets are preserved as formulas, not converted to cached values.
    """
    wb = openpyxl.load_workbook(ledger_path)
    ws = wb["Transactions"]
    for c in changes:
        ws.cell(c["row"], cat_col_1idx).value = c["new_cat"]
    wb.save(ledger_path)
    wb.close()


# ── report ────────────────────────────────────────────────────────────────────

def _print_report(changes: list, stats: dict, applied: bool, ledger_path: str) -> None:
    mode = "APPLIED" if applied else "DRY-RUN"
    verb = "made" if applied else "found"

    print(f"\n{'='*W}")
    print(f"recategorize_ledger.py  —  {mode}")
    print(f"{'='*W}")
    print(f"  Ledger  : {ledger_path}")
    print(f"  Rows scanned           : {stats['rows_total']}")
    print(f"  Protected (skipped)    : {stats['skipped_protected']}"
          f"   ← IncludeInNet=False or One-Off category")
    print(f"  New=excluded (skipped) : {stats['skipped_new_excluded']}"
          f"   ← new rule says Internal Transfer; needs manual review")
    print(f"  PFC-guard (skipped)    : {stats['skipped_pfc_guard']}"
          f"   ← would downgrade PFC-sourced category to Other-Uncategorized")
    print(f"  Changes {verb:5}         : {len(changes)}")

    if changes:
        # Summary by (old → new) pair
        pairs = Counter((c["old_cat"], c["new_cat"]) for c in changes)
        print(f"\n{'─'*W}")
        print(f"  Change summary (old category → new category):")
        print(f"{'─'*W}")
        for (old, new), count in sorted(pairs.items(), key=lambda x: -x[1]):
            print(f"  {count:>3}x  {old:<44} →  {new}")

        print(f"\n{'─'*W}")
        print(f"  {'Date':<12} {'Type':<8} {'Amount':>10}   {'Old Category':<38} →  New Category")
        print(f"{'─'*W}")
        for c in sorted(changes, key=lambda x: x["date"]):
            print(f"  {c['date']:<12} {c['type']:<8} ${c['amount']:>9,.2f}"
                  f"   {c['old_cat']:<38} →  {c['new_cat']}")
            print(f"  {'':12}  {c['desc']}")
    else:
        print(f"\n  ✓ No changes needed — all rows already match current rules.")

    # Raw income/expense totals — always identical before/after because we
    # only change Category, never Type or Amount.
    inc_b = stats["income_before"]
    exp_b = stats["expense_before"]
    print(f"\n{'─'*W}")
    print(f"  Income/Expense totals (Type+Amount only — unchanged by recategorization):")
    print(f"    Income :  ${inc_b:>13,.2f}")
    print(f"    Expense:  ${exp_b:>13,.2f}")
    print(f"    Net    :  ${inc_b - exp_b:>13,.2f}")
    print(f"{'='*W}\n")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Re-apply current spending rules to all Transactions rows (Category column only).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Default is dry-run. Pass --apply to write changes to the ledger.",
    )
    parser.add_argument(
        "--apply", action="store_true",
        help="Write changes to the ledger (a timestamped backup is made first).",
    )
    parser.add_argument(
        "--ledger", metavar="PATH",
        help="Path to the ledger .xlsx file. Overrides the default (SPENDING_LEDGER_FILE_PATH / resolve_ledger_path).",
    )
    args = parser.parse_args()

    if args.ledger:
        ledger_path = args.ledger
        if not os.path.exists(ledger_path):
            sys.exit(f"Error: Ledger not found at {ledger_path}")
    else:
        ledger_path, _ = resolve_ledger_path()
        if not ledger_path:
            sys.exit("Error: SPENDING_LEDGER_FILE_PATH not configured.")
        if not os.path.exists(ledger_path):
            sys.exit(f"Error: Ledger not found at {ledger_path}")

    if not os.path.exists(RULES_PATH):
        sys.exit(f"Error: spending_rules.json not found at {RULES_PATH}")

    rules = load_rules(RULES_PATH)

    print(f"Rules  : {len(rules)} loaded from {os.path.basename(RULES_PATH)}")
    print(f"Ledger : {ledger_path}")
    print(f"Mode   : {'--apply (WILL WRITE)' if args.apply else 'dry-run (read-only)'}")

    # Read category column index (1-based) from the header before any writes.
    wb_hdr = openpyxl.load_workbook(ledger_path, read_only=True, data_only=True)
    raw_header = [c.value for c in next(wb_hdr["Transactions"].iter_rows(min_row=1, max_row=1))]
    wb_hdr.close()
    try:
        cat_col_1idx = next(i + 1 for i, name in enumerate(raw_header) if name == "Category")
    except StopIteration:
        sys.exit("Error: 'Category' column not found in Transactions header.")

    # Always scan first (read-only)
    print("\nScanning…")
    changes, stats = _scan(ledger_path, rules)
    _print_report(changes, stats, applied=False, ledger_path=ledger_path)

    if not args.apply:
        if changes:
            print(f"Run with --apply to apply {len(changes)} change(s).")
        return

    if not changes:
        print("Nothing to apply.")
        return

    # Make backup before touching anything
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_path = ledger_path.replace(".xlsx", f"_BACKUP_{stamp}.xlsx")
    shutil.copy2(ledger_path, backup_path)
    print(f"Backup : {backup_path}")

    # Write changes
    _apply(ledger_path, changes, cat_col_1idx)
    print(f"Applied {len(changes)} category update(s).\n")

    # Verify: re-scan must find zero remaining changes and identical row count
    print("Verifying…")
    changes_after, stats_after = _scan(ledger_path, rules)

    rows_before = stats["rows_total"]
    rows_after  = stats_after["rows_total"]
    inc_before  = round(stats["income_before"],  2)
    exp_before  = round(stats["expense_before"], 2)
    inc_after   = round(stats_after["income_before"],  2)
    exp_after   = round(stats_after["expense_before"], 2)

    print(f"\n{'='*W}")
    print(f"POST-APPLY VERIFICATION")
    print(f"{'='*W}")
    print(f"  Row count       : {rows_before} → {rows_after}"
          f"  {'✓ unchanged' if rows_before == rows_after else '✗ CHANGED — BUG'}")
    print(f"  Income total    : ${inc_before:>13,.2f} → ${inc_after:>13,.2f}"
          f"  {'✓ unchanged' if inc_before == inc_after else '✗ CHANGED (unexpected)'}")
    print(f"  Expense total   : ${exp_before:>13,.2f} → ${exp_after:>13,.2f}"
          f"  {'✓ unchanged' if exp_before == exp_after else '✗ CHANGED (unexpected)'}")
    print(f"  Residual diffs  : {len(changes_after)}"
          f"  {'✓ zero' if not changes_after else '✗ NON-ZERO — BUG'}")

    if changes_after:
        print("\n  ✗ BUG: the following rows still differ after apply:")
        for c in changes_after:
            print(f"    row {c['row']}  {c['date']}  {c['old_cat']!r} → {c['new_cat']!r}")

    if rows_before != rows_after or changes_after:
        print(f"\n  ✗ Verification FAILED. Backup preserved at: {backup_path}")
        sys.exit(1)

    print(f"{'='*W}")
    print(f"\n  Backup : {backup_path}")
    print(f"  Done.\n")


if __name__ == "__main__":
    main()
