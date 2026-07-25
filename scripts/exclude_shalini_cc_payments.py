#!/usr/bin/env python3
"""Neutralize the historical lump-sum payments to Shalini's BofA VISA (x3070)
so they don't double-count once her card's itemized purchases are synced.

Before x3070 was synced item-level, each monthly bill-pay from checking x5799
to ACCT# 3070 was counted as an expense ("Other Credit Card Expenses") — a
proxy for her spend. Now that the real purchases are imported, those lump
payments must be excluded (they're internal transfers, not spend).

This targets BOTH legs of the internal transfer that leaked into totals:
  • checking side: "Online Scheduled Payment to ACCT# 3070" (the old proxy
    lumps, written as "Other Credit Card Expenses"), and
  • card side: "FROM CHK 5799" (e.g. "ONLINE/MOBILE RECURRING FROM CHK 5799"),
    which the Plaid backfill booked as fake income (~$2,020).
For each match it:
  • sets Category   -> "Credit Card Payment"  (the excluded-transfer bucket)
  • sets IncludeInNet -> False                 (drops it from Income/Expense/Net)
The row stays visible in the Transactions sheet as an audit trail.

Future occurrences never reach the sheet: spending_rules.json now maps both
"Online Scheduled Payment to ACCT# 3070" and "FROM CHK 5799" to
"Credit Card Payment", which the sync excludes before writing. This one-time
fix only touches rows written before those rules were in place.

Usage:
    python scripts/exclude_shalini_cc_payments.py --dry-run
    python scripts/exclude_shalini_cc_payments.py
    python scripts/exclude_shalini_cc_payments.py --from-date 2026-05-01   # only rows on/after
"""

import os
import shutil
import sys
from datetime import datetime, date

import openpyxl
from dotenv import load_dotenv

load_dotenv()

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LEDGER = os.environ.get("SPENDING_LEDGER_FILE_PATH", os.path.join(ROOT, "cashflow-tracker.xlsx"))

KEYWORDS = ["Online Scheduled Payment to ACCT# 3070", "FROM CHK 5799"]
NEW_CATEGORY = "Credit Card Payment"


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def main():
    dry = "--dry-run" in sys.argv
    from_date = None
    if "--from-date" in sys.argv:
        from_date = date.fromisoformat(sys.argv[sys.argv.index("--from-date") + 1])

    wb = openpyxl.load_workbook(LEDGER)
    ws = wb["Transactions"]
    header = [c.value for c in ws[1]]
    di = header.index("Date")
    desc_i = header.index("Description")
    cat_i = header.index("Category")
    amt_i = header.index("Amount")
    inc_i = header.index("IncludeInNet")

    matches = []
    for row in ws.iter_rows(min_row=2):
        desc = str(row[desc_i].value or "")
        if not any(kw in desc for kw in KEYWORDS):
            continue
        d = _parse_date(row[di].value)
        if from_date and (d is None or d < from_date):
            continue
        matches.append(row)

    print(f"Ledger: {LEDGER}")
    print(f"Matched {len(matches)} row(s) for {KEYWORDS}"
          + (f" on/after {from_date}" if from_date else "") + ":")
    total = 0.0
    for row in matches:
        already = (row[cat_i].value == NEW_CATEGORY and not bool(row[inc_i].value))
        amt = float(row[amt_i].value or 0)
        total += amt
        flag = "  [already excluded]" if already else ""
        print(f"  {str(row[di].value)[:10]:10} | {str(row[cat_i].value):26} | "
              f"IncludeInNet={row[inc_i].value} | ${amt:,.2f}{flag}")

    if dry:
        print(f"\nDRY RUN — no changes written. Would exclude ${total:,.2f} across "
              f"{len(matches)} row(s).")
        wb.close()
        return

    if not matches:
        print("Nothing to change.")
        wb.close()
        return

    backup = LEDGER.replace(".xlsx", f"_BACKUP_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx")
    shutil.copy2(LEDGER, backup)

    changed = 0
    for row in matches:
        row[cat_i].value = NEW_CATEGORY
        row[inc_i].value = False
        changed += 1

    wb.save(LEDGER)
    print(f"\nExcluded {changed} row(s) (${total:,.2f} removed from expense totals). "
          f"Backup: {os.path.basename(backup)}")
    print("Next: 'push' from the agent CLI to sync Drive.")


if __name__ == "__main__":
    main()
