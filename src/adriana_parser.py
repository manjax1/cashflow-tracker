import io
import csv
import re
from datetime import datetime, timedelta, date as _date

import openpyxl
from googleapiclient.http import MediaIoBaseDownload

ADRIANA_FOLDER_ID = "1hP-wFMW_Dk66D-jfo3k-G0M1w2DnzyIr"

# Ordered: 1867A must precede 1867 to avoid substring false-match.
# Match requires BOTH substrings (number + street) to be present.
_PROPERTY_RULES = [
    ("807",   "center",  "Manteca-807"),
    ("809",   "center",  "Manteca-809"),
    ("2516",  "mission", "Stockton-2516"),
    ("2528",  "mission", "Stockton-2528"),
    ("1867a", "elmwood", "Stockton-1867A"),
    ("1867",  "elmwood", "Stockton-1867"),
]

_MONTHS = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june",
     "july", "august", "september", "october", "november", "december"], 1
)}

_EXCEL_EPOCH = datetime(1899, 12, 30)

_SKIP_PAYEES = {"beginning balance", "ending balance", "reserves"}


def _map_property(raw: str) -> str | None:
    if not raw:
        return None
    lower = raw.lower()
    for num, street, label in _PROPERTY_RULES:
        if num in lower and street in lower:
            return label
    return None


def _meta_flag_is_set(wb, key: str) -> bool:
    if "_Meta" not in wb.sheetnames:
        return False
    ws = wb["_Meta"]
    for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
        if row and str(row[0]) == key:
            return bool(row[1])
    return False


def _parse_xlsx_date(raw) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, _date):  # catches both datetime and date
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, (int, float)):
        try:
            return (_EXCEL_EPOCH + timedelta(days=int(raw))).strftime("%Y-%m-%d")
        except Exception:
            return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(str(raw).strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _dedup_key(year_month: str, property_short: str, date_str: str, amount: float) -> str:
    return f"adriana:{year_month}:{property_short}:{date_str}:{amount:.2f}"


def list_unprocessed_adriana_files(service, wb) -> list[dict]:
    pattern = re.compile(
        r"Adriana Managed Properties Ledger\s*-\s*(\w+)\s+(\d{4})",
        re.IGNORECASE,
    )
    resp = service.files().list(
        q=f"'{ADRIANA_FOLDER_ID}' in parents and trashed = false",
        fields="files(id, name, mimeType)",
    ).execute()

    result = []
    for f in resp.get("files", []):
        m = pattern.search(f["name"])
        if not m:
            continue
        month_num = _MONTHS.get(m.group(1).lower())
        if not month_num:
            print(f"⚠️  Adriana: unrecognized month in filename: {f['name']}")
            continue
        year = int(m.group(2))
        ym = f"{year}-{month_num:02d}"
        if _meta_flag_is_set(wb, f"adriana_processed:{ym}"):
            continue
        result.append({
            "file_id":   f["id"],
            "name":      f["name"],
            "year":      year,
            "month":     month_num,
            "mime_type": f["mimeType"],
        })
    return result


def _cell(row, idx) -> str:
    """Safely extract a cell value as a stripped string; returns '' on missing/None."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _parse_xlsx(content: bytes, file_meta: dict) -> list[dict]:
    year_month = f"{file_meta['year']}-{file_meta['month']:02d}"
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    transactions = []

    for ws in wb.worksheets:
        # Locate the header row (search first 20 rows)
        header_idx = None
        col_map: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            row_strs = [str(c).strip().lower() if c is not None else "" for c in row]
            if "date" in row_strs and "property" in row_strs and "income" in row_strs:
                header_idx = r_idx
                for c_idx, cell_val in enumerate(row):
                    if cell_val is not None:
                        col_map[str(cell_val).strip().lower()] = c_idx
                break
        if header_idx is None:
            continue

        date_col    = col_map.get("date")
        prop_col    = col_map.get("property")
        payee_col   = col_map.get("payee/payer")
        income_col  = col_map.get("income")
        expense_col = col_map.get("expense")
        notes_col   = col_map.get("notes")

        if date_col is None or prop_col is None:
            continue

        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            if not any(c for c in row if c is not None):
                continue

            # Date
            date_raw = row[date_col] if date_col < len(row) else None
            date_str = _parse_xlsx_date(date_raw)
            if not date_str:
                if date_raw is not None:
                    print(f"⚠️  Adriana XLSX '{file_meta['name']}': bad date {date_raw!r} — skipping row")
                continue

            # Property
            prop_raw = _cell(row, prop_col)
            prop_label = _map_property(prop_raw)
            if not prop_label:
                if prop_raw:
                    print(f"⚠️  Adriana XLSX: unrecognized property {prop_raw!r} — skipping row")
                continue

            # Payee and notes
            payee_raw = _cell(row, payee_col)
            payee_low = payee_raw.lower()
            notes_low = _cell(row, notes_col).lower()

            if payee_low in _SKIP_PAYEES:
                continue
            if "manjunath" in payee_low:
                continue  # landlord payment — already captured in Plaid

            # Amounts
            def _amt(idx) -> float:
                v = row[idx] if idx is not None and idx < len(row) else None
                if v is None or v == "":
                    return 0.0
                try:
                    return abs(float(v))
                except (ValueError, TypeError):
                    return 0.0

            income_amt  = _amt(income_col)
            expense_amt = _amt(expense_col)

            # Classify
            if "jpeter" in payee_low or "management fee" in notes_low:
                category = "Rental - Management Fee"
                tx_type  = "Expense"
                amount   = expense_amt if expense_amt > 0 else income_amt
            elif income_amt > 0:
                category = "Rental - Income"
                tx_type  = "Income"
                amount   = income_amt
            elif expense_amt > 0:
                category = "Rental - Maintenance"
                tx_type  = "Expense"
                amount   = expense_amt
            else:
                continue

            if amount <= 0:
                continue

            transactions.append({
                "date":             date_str,
                "name":             payee_raw or "(unknown)",
                "account_label":    prop_label,
                "category":         category,
                "type":             tx_type,
                "amount":           round(amount, 2),
                "transaction_id":   _dedup_key(year_month, prop_label, date_str, amount),
                "exclude_from_net": False,
            })

    return transactions


def _parse_csv(content: bytes, file_meta: dict) -> list[dict]:
    year_month = f"{file_meta['year']}-{file_meta['month']:02d}"
    text = content.decode("utf-8-sig")  # handle Windows BOM
    transactions = []

    for row in csv.DictReader(io.StringIO(text)):
        # Property
        prop_raw   = row.get("Property Name", "").strip()
        prop_label = _map_property(prop_raw)
        if not prop_label:
            if prop_raw:
                print(f"⚠️  Adriana CSV: unrecognized property {prop_raw!r} — skipping row")
            continue

        # Date ("M/D/YY")
        date_raw = row.get("Date", "").strip()
        try:
            date_str = datetime.strptime(date_raw, "%m/%d/%y").strftime("%Y-%m-%d")
        except ValueError:
            print(f"⚠️  Adriana CSV '{file_meta['name']}': bad date {date_raw!r} — skipping row")
            continue

        # Category → type
        category_raw = row.get("Category", "").strip()
        if category_raw == "Rents Received":
            category = "Rental - Income"
            tx_type  = "Income"
        elif category_raw == "Management Fees":
            category = "Rental - Management Fee"
            tx_type  = "Expense"
        else:
            continue  # unknown category — skip per spec

        # Amount
        amount_raw = row.get("Amount", "").strip().lstrip("$").replace(",", "")
        try:
            amount = abs(float(amount_raw))
        except ValueError:
            print(f"⚠️  Adriana CSV: bad amount {amount_raw!r} — skipping row")
            continue
        if amount <= 0:
            continue

        name = row.get("Transaction Name", "").strip() or prop_raw

        transactions.append({
            "date":             date_str,
            "name":             name,
            "account_label":    prop_label,
            "category":         category,
            "type":             tx_type,
            "amount":           round(amount, 2),
            "transaction_id":   _dedup_key(year_month, prop_label, date_str, amount),
            "exclude_from_net": False,
        })

    return transactions


def parse_adriana_file(service, file_meta: dict) -> list[dict]:
    buf = io.BytesIO()
    req = service.files().get_media(fileId=file_meta["file_id"])
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    content = buf.getvalue()

    mime = file_meta["mime_type"]
    if "spreadsheetml" in mime:
        return _parse_xlsx(content, file_meta)
    elif "csv" in mime or mime == "text/plain":
        return _parse_csv(content, file_meta)
    else:
        raise ValueError(f"Unsupported MIME type for Adriana file: {mime!r}")
