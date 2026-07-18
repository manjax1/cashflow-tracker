"""Tool schemas + dispatch. Read tools execute immediately.
Action tools only create proposals; execution happens in cli.py after
explicit user approval (the gate lives in code, not in the prompt)."""

import json
import os
import shutil
import time
from datetime import datetime

import openpyxl

from . import ledger

REPO_ROOT = ledger.REPO_ROOT
OUTBOX_DIR = os.path.join(REPO_ROOT, "outbox")
AUDIT_LOG = os.path.join(REPO_ROOT, "logs", "agent_audit.jsonl")

TOOLS = [
    {
        "name": "query_transactions",
        "description": "Query ledger transactions with filters. Returns JSON rows sorted newest-first. Use 'search' for description keyword matching.",
        "input_schema": {
            "type": "object",
            "properties": {
                "start_date": {"type": "string", "description": "ISO date YYYY-MM-DD"},
                "end_date": {"type": "string", "description": "ISO date YYYY-MM-DD"},
                "category": {"type": "string", "description": "Category substring filter (case-insensitive), e.g. 'Rental' matches all Rental - * categories; 'Uncategorized' matches 'Other - Uncategorized'"},
                "account": {"type": "string", "description": "e.g. 'Checking' or 'Credit Card'"},
                "tx_type": {"type": "string", "enum": ["Income", "Expense"]},
                "min_amount": {"type": "number"},
                "max_amount": {"type": "number"},
                "search": {"type": "string", "description": "Substring match on description"},
                "limit": {"type": "integer", "default": 100},
            },
            "required": ["start_date", "end_date"],
        },
    },
    {
        "name": "get_cashflow_summary",
        "description": "Computed aggregates (income, expense, net) for a period, grouped. THIS tool does the math — never compute totals yourself. Includes a Rental - * rollup.",
        "input_schema": {
            "type": "object",
            "properties": {
                "start_date": {"type": "string"},
                "end_date": {"type": "string"},
                "group_by": {"type": "string", "enum": ["category", "month", "account", "type"], "default": "category"},
                "category": {"type": "string", "description": "Optional category prefix filter"},
            },
            "required": ["start_date", "end_date"],
        },
    },
    {
        "name": "get_trends",
        "description": "Deterministic time series for a metric with period-over-period deltas and trailing average. Use for any 'trend', 'over time', or 'growing/shrinking' question. Interpret the series; do not recompute it.",
        "input_schema": {
            "type": "object",
            "properties": {
                "metric": {"type": "string", "enum": ["income", "expenses", "net"], "default": "net"},
                "granularity": {"type": "string", "enum": ["month", "quarter"], "default": "month"},
                "lookback_periods": {"type": "integer", "default": 6},
                "category": {"type": "string", "description": "Optional category prefix filter"},
            },
        },
    },
    {
        "name": "list_categories",
        "description": "All ledger categories with counts, totals, and date ranges. Call this first when the user names a category colloquially.",
        "input_schema": {"type": "object", "properties": {}},
    },
    {
        "name": "get_category_rules",
        "description": "The categorization rules engine's keyword rules. Use to explain WHY a transaction was categorized a certain way, or before proposing a new rule.",
        "input_schema": {
            "type": "object",
            "properties": {"keyword_filter": {"type": "string", "description": "Filter rules by substring"}},
        },
    },
    {
        "name": "find_anomalies",
        "description": "Deterministic anomaly checks: duplicates, unusually large transactions, missing expected rental income, category spend spikes, uncategorized backlog. Your job is to explain and prioritize the findings, not detect them.",
        "input_schema": {
            "type": "object",
            "properties": {"lookback_days": {"type": "integer", "default": 30}},
        },
    },
    {
        "name": "draft_email",
        "description": "PROPOSAL ONLY: draft an email (e.g., late-rent reminder, vendor query). The draft is saved to an outbox and shown to the user for approval — it is NEVER sent automatically. Tell the user a draft was created for their review.",
        "input_schema": {
            "type": "object",
            "properties": {
                "recipient_hint": {"type": "string", "description": "Who this is for, e.g. 'tenant at Mulholland Drive'"},
                "subject": {"type": "string"},
                "body": {"type": "string"},
            },
            "required": ["recipient_hint", "subject", "body"],
        },
    },
    {
        "name": "recategorize_batch",
        "description": "PROPOSAL ONLY: propose recategorizing MANY transactions in one call. Preferred over recategorize_transaction for >3 changes. Each item needs the SourceRef from query_transactions and an EXACT existing category name from list_categories. The user reviews every item individually (approve/edit/discard) before anything is modified.",
        "input_schema": {
            "type": "object",
            "properties": {
                "items": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "source_ref": {"type": "string"},
                            "new_category": {"type": "string"},
                            "reason": {"type": "string"},
                            "rule_keyword": {"type": "string", "description": "Optional: add a spending_rules.json rule for recurring merchants"},
                        },
                        "required": ["source_ref", "new_category"],
                    },
                }
            },
            "required": ["items"],
        },
    },
    {
        "name": "recategorize_transaction",
        "description": "PROPOSAL ONLY: propose changing a transaction's category (and optionally add a keyword rule so future transactions match). Requires the transaction's SourceRef from query_transactions. The user approves before the ledger is modified.",
        "input_schema": {
            "type": "object",
            "properties": {
                "source_ref": {"type": "string", "description": "SourceRef of the transaction"},
                "new_category": {"type": "string"},
                "reason": {"type": "string"},
                "rule_keyword": {"type": "string", "description": "Optional: keyword for a new spending_rules.json rule"},
            },
            "required": ["source_ref", "new_category", "reason"],
        },
    },
]

ACTION_TOOLS = {"draft_email", "recategorize_transaction", "recategorize_batch"}

# Proposals accumulated during a run; cli.py drains and gates them.
PENDING = []


def audit(event, detail):
    os.makedirs(os.path.dirname(AUDIT_LOG), exist_ok=True)
    with open(AUDIT_LOG, "a") as f:
        f.write(json.dumps({"ts": datetime.now().isoformat(timespec="seconds"),
                            "event": event, "detail": detail}) + "\n")


def dispatch(name, args):
    """Execute a read tool, or register a proposal for an action tool."""
    audit("tool_call", {"tool": name, "args": args})
    if name in ACTION_TOOLS:
        proposal = {"id": f"p{int(time.time()*1000)}", "tool": name, "args": args}
        PENDING.append(proposal)
        return {"status": "proposal_created", "proposal_id": proposal["id"],
                "note": "Awaiting user approval. Nothing has been sent or modified."}
    fn = {
        "query_transactions": ledger.query_transactions,
        "get_cashflow_summary": ledger.get_cashflow_summary,
        "get_trends": ledger.get_trends,
        "list_categories": lambda: ledger.list_categories(),
        "get_category_rules": ledger.get_category_rules,
        "find_anomalies": ledger.find_anomalies,
    }[name]
    try:
        return fn(**args) if args else fn()
    except Exception as e:  # return errors to the model so it can adapt
        return {"error": f"{type(e).__name__}: {e}"}


# ------------------------- approved executions -------------------------

def execute_draft_email(args):
    os.makedirs(OUTBOX_DIR, exist_ok=True)
    path = os.path.join(OUTBOX_DIR, f"draft_{datetime.now():%Y%m%d_%H%M%S}.json")
    with open(path, "w") as f:
        json.dump(args, f, indent=2)
    audit("email_draft_approved", {"path": path, "subject": args["subject"]})
    return f"Draft saved to {path}. (Wire to email_notifier.py to actually send.)"


def execute_recategorize(args):
    # Backup first — consistent with existing BACKUP_* convention.
    backup = ledger.LEDGER_PATH.replace(
        ".xlsx", f"_BACKUP_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx")
    shutil.copy2(ledger.LEDGER_PATH, backup)

    wb = openpyxl.load_workbook(ledger.LEDGER_PATH)
    ws = wb["Transactions"]
    header = [c.value for c in ws[1]]
    ref_col = header.index("SourceRef") + 1
    cat_col = header.index("Category") + 1
    hit = None
    for row in ws.iter_rows(min_row=2):
        if str(row[ref_col - 1].value) == args["source_ref"]:
            old = row[cat_col - 1].value
            row[cat_col - 1].value = args["new_category"]
            hit = old
            break
    if hit is None:
        return f"SourceRef {args['source_ref']} not found; ledger unchanged (backup at {backup})."
    wb.save(ledger.LEDGER_PATH)
    ledger._cache["mtime"] = None  # invalidate cache

    rule_msg = ""
    if args.get("rule_keyword"):
        with open(ledger.RULES_PATH) as f:
            rules = json.load(f)
        rules.append({"keyword": args["rule_keyword"], "category": args["new_category"],
                      "note": f"Added by cashflow agent: {args['reason']}"})
        with open(ledger.RULES_PATH, "w") as f:
            json.dump(rules, f, indent=2)
        rule_msg = f" Rule added for keyword '{args['rule_keyword']}'."

    audit("recategorize_approved", {**args, "old_category": hit, "backup": backup})
    return (f"Recategorized {args['source_ref']}: '{hit}' -> '{args['new_category']}'."
            f"{rule_msg} Backup: {os.path.basename(backup)}")


def execute_recategorize_batch(args):
    """Apply many approved recategorizations with ONE backup and ONE save."""
    items = args["items"]
    if not items:
        return "No approved items; ledger unchanged."
    backup = ledger.LEDGER_PATH.replace(
        ".xlsx", f"_BACKUP_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx")
    shutil.copy2(ledger.LEDGER_PATH, backup)

    wb = openpyxl.load_workbook(ledger.LEDGER_PATH)
    ws = wb["Transactions"]
    header = [c.value for c in ws[1]]
    ref_col = header.index("SourceRef")
    cat_col = header.index("Category")
    by_ref = {}
    for row in ws.iter_rows(min_row=2):
        by_ref[str(row[ref_col].value)] = row
    applied, missing, new_rules = [], [], []
    for it in items:
        row = by_ref.get(it["source_ref"])
        if row is None:
            missing.append(it["source_ref"])
            continue
        old = row[cat_col].value
        row[cat_col].value = it["new_category"]
        applied.append({**it, "old_category": old})
        if it.get("rule_keyword"):
            new_rules.append({"keyword": it["rule_keyword"],
                              "category": it["new_category"],
                              "note": f"Added by cashflow agent: {it.get('reason', 'batch recategorization')}"})
    wb.save(ledger.LEDGER_PATH)
    ledger._cache["mtime"] = None

    if new_rules:
        with open(ledger.RULES_PATH) as f:
            rules = json.load(f)
        existing = {r["keyword"].lower() for r in rules}
        new_rules = [r for r in new_rules if r["keyword"].lower() not in existing]
        rules.extend(new_rules)
        with open(ledger.RULES_PATH, "w") as f:
            json.dump(rules, f, indent=2)

    audit("recategorize_batch_approved",
          {"applied": applied, "missing": missing,
           "rules_added": [r["keyword"] for r in new_rules], "backup": backup})
    msg = f"Applied {len(applied)} recategorizations."
    if new_rules:
        msg += f" Added {len(new_rules)} rules."
    if missing:
        msg += f" {len(missing)} SourceRefs not found: {missing[:5]}"
    return msg + f" Backup: {os.path.basename(backup)}"


EXECUTORS = {"draft_email": execute_draft_email,
             "recategorize_transaction": execute_recategorize,
             "recategorize_batch": execute_recategorize_batch}
