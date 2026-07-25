import os
import sys
import base64
import argparse
from datetime import date, timedelta, datetime
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv

load_dotenv()

from utils import clean_env, resolve_ledger_path
from plaid_client import PlaidClient
from filters import load_rules, categorize_batch
from openpyxl import load_workbook
from ledger_writer import write_spending_ledger, get_last_snapshot_month, set_last_snapshot_month, set_meta_flag
from email_notifier import send_sync_summary
from drive_sync import download_ledger, upload_ledger, get_drive_service

RULES_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "spending_rules.json")


def _resolve_ledger_path() -> tuple[str, bool]:
    return resolve_ledger_path()


def _load_rules_with_fallback() -> list:
    """Load categorization rules. Priority:
    1. Google Drive (RULES_DRIVE_FILE_ID) — updated via scripts/push_rules_to_drive.py,
       no redeploy, single source of truth alongside the ledger.
    2. RULES_JSON env var (legacy Railway).
    3. Local spending_rules.json file.
    """
    import json
    from filters import load_rules

    rules_drive_id = clean_env(os.getenv("RULES_DRIVE_FILE_ID"), "RULES_DRIVE_FILE_ID")
    if rules_drive_id:
        try:
            download_ledger(rules_drive_id, RULES_PATH)   # generic Drive file download
            rules = load_rules(RULES_PATH)
            print(f"✅ Loaded {len(rules)} rules from Drive")
            return rules
        except Exception as e:
            print(f"⚠️  Drive rules load failed: {e} — falling back")

    rules_json_env = clean_env(os.getenv("RULES_JSON"), "RULES_JSON")
    if rules_json_env:
        try:
            rules = json.loads(rules_json_env)
            print(f"✅ Loaded {len(rules)} rules from RULES_JSON env var")
            from filters import load_rules as _sort_rules
            return sorted(rules, key=lambda r: len(r["keyword"]), reverse=True)
        except Exception as e:
            print(f"⚠️  RULES_JSON parse failed: {e} — falling back to file")

    if os.path.exists(RULES_PATH):
        from filters import load_rules
        rules = load_rules(RULES_PATH)
        print(f"✅ Loaded {len(rules)} rules from {RULES_PATH}")
        return rules

    print("⚠️  No spending_rules.json and no RULES_JSON env var — using Plaid categories only.")
    return []


def run_sync(from_date: date = None, to_date: date = None) -> dict:
    is_cloud = bool(os.getenv("RAILWAY_ENVIRONMENT"))
    plaid_env = clean_env(os.getenv("PLAID_ENV", "production"), "PLAID_ENV")
    resend_set = bool(clean_env(os.getenv("RESEND_API_KEY"), "RESEND_API_KEY"))
    token_set = bool(clean_env(os.getenv("PLAID_ACCESS_TOKEN"), "PLAID_ACCESS_TOKEN"))
    drive_set = bool(clean_env(os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON"), "GOOGLE_SERVICE_ACCOUNT_JSON"))
    print(f"[spending-tracker] env={plaid_env} cloud={is_cloud} resend={resend_set} token={token_set} drive={drive_set}")

    ledger_path, is_cloud = _resolve_ledger_path()
    file_id = clean_env(os.getenv("GOOGLE_DRIVE_FILE_ID"), "GOOGLE_DRIVE_FILE_ID")

    if is_cloud and file_id:
        try:
            download_ledger(file_id, ledger_path)
        except Exception:
            print("⚠️  Could not download existing ledger — will create fresh file.")

    client = PlaidClient()

    end = to_date if to_date else date.today()
    start = from_date if from_date else end - timedelta(days=7)
    print(f"Fetching transactions {start} → {end}")

    # ── Plaid items ──────────────────────────────────────────────────────
    # Each item is a separate Plaid login.
    #   • primary — your BofA login. Many accounts; filtered by an EXCLUDE
    #     list (everything is kept except the masks below).
    #   • additional items (e.g. Shalini's BofA login) use an INCLUDE list so
    #     we pull ONLY the named account(s). This is what prevents double-
    #     counting anything the primary item already covers — notably the
    #     joint checking x5799, which lives under both logins.
    #
    # 6450 (Tanusha credit card) is intentionally NOT excluded — tracked.
    PRIMARY_EXCLUDED_MASKS = {
        "4719": "Prateek checking (son)",
        "0043": "Tanusha checking (daughter)",
        "5663": "Adv Relationship Banking (2nd checking, no household activity)",
        "8305": "Primary mortgage loan account (payment already captured via checking debit)",
    }

    items = [{
        "name": "primary",
        "token": clean_env(os.getenv("PLAID_ACCESS_TOKEN"), "PLAID_ACCESS_TOKEN"),
        "mode": "exclude",
        "masks": set(PRIMARY_EXCLUDED_MASKS),
        "label_override": None,
        "required": True,
    }]

    # Shalini's BofA login — sync ONLY her Premium Rewards Visa (x3070).
    # Her joint checking (x5799) is already captured via the primary item,
    # so the include-list deliberately omits it to avoid double-counting.
    shalini_token = clean_env(os.getenv("PLAID_ACCESS_TOKEN_SHALINI"), "PLAID_ACCESS_TOKEN_SHALINI")
    if shalini_token:
        items.append({
            "name": "shalini",
            "token": shalini_token,
            "mode": "include",
            "masks": {"3070"},
            "label_override": "Shalini BoA VISA",
            "required": False,
        })

    raw_transactions: list = []
    account_map: dict = {}
    excluded_by_account: dict[str, list] = {mask: [] for mask in PRIMARY_EXCLUDED_MASKS}

    for item in items:
        token = item["token"]
        # Verify the token. The primary item is mandatory (locally we fall
        # back to the interactive link flow); additional items are best-effort
        # so an expired secondary login never blocks the household sync.
        if not token or not client.verify_access_token(token):
            if item["required"]:
                if is_cloud:
                    raise RuntimeError("Plaid access token invalid or missing. Run link flow locally first.")
                from link_flow import run_link_flow
                run_link_flow(client)
                token = clean_env(os.getenv("PLAID_ACCESS_TOKEN"), "PLAID_ACCESS_TOKEN")
            else:
                print(f"⚠️  Plaid item '{item['name']}' token missing/invalid — skipping "
                      f"(re-link with: python src/link_item.py {item['name']}).")
                continue

        item_txns = client.get_transactions(token, start, end)
        item_accounts = client.get_accounts(token)

        mode, masks, override = item["mode"], item["masks"], item.get("label_override")
        mask_by_id = {a["account_id"]: a.get("mask", "") for a in item_accounts}

        kept_ids = set()
        for a in item_accounts:
            mask = a.get("mask", "")
            keep = (mask in masks) if mode == "include" else (mask not in masks)
            if keep:
                account_map[a["account_id"]] = override or _account_label(a)
                kept_ids.add(a["account_id"])

        kept = dropped = 0
        for tx in item_txns:
            aid = tx.get("account_id", "")
            if aid in kept_ids:
                raw_transactions.append(tx)
                kept += 1
            else:
                dropped += 1
                # Primary-item exclusions are reported in the sync summary.
                mask = mask_by_id.get(aid, "")
                if item["name"] == "primary" and mask in excluded_by_account:
                    excluded_by_account[mask].append(tx)

        if item["name"] != "primary":
            print(f"[item:{item['name']}] {kept} txns kept "
                  f"(masks {sorted(masks)}), {dropped} from other accounts skipped")

    rules = _load_rules_with_fallback()

    included, excluded = categorize_batch(raw_transactions, rules, account_map)

    # Warn if any excluded transaction's name contains rental-related signals —
    # guards against broad exclusion rules accidentally suppressing rental income
    # or expenses (the same class of bug as the Hari Vasantapu keyword-shadowing
    # fix). CC payments are always intentional and excluded from this count.
    _RENTAL_NAME_SIGNALS = ("rent", "tenant", "lease")
    excluded_rental_count = sum(
        1 for tx in excluded
        if tx.get("category") != "Credit Card Payment"
        and any(sig in (tx.get("name") or "").lower() for sig in _RENTAL_NAME_SIGNALS)
    )

    result = write_spending_ledger(ledger_path, included)
    added, skipped = result["added"], result["skipped"]
    new_transactions = result.get("new_transactions", [])

    # Build summary
    category_totals: dict[str, float] = defaultdict(float)
    for tx in included:
        if tx.get("type") == "Expense":
            category_totals[tx["category"]] += tx.get("amount", 0.0)

    sorted_cats = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)
    total_spend = sum(v for _, v in sorted_cats)
    top_category = sorted_cats[0][0] if sorted_cats else "N/A"

    summary = {
        "date": str(end),
        "added": added,
        "skipped": skipped,
        "tx_count": added,
        "new_transactions": new_transactions,
        "total_spend": total_spend,
        "top_category": top_category,
        "top_categories": [{"category": c, "amount": a} for c, a in sorted_cats[:5]],
        "transactions": [
            {
                "date": tx.get("date"),
                "name": tx.get("name"),
                "account_label": tx.get("account_label"),
                "category": tx.get("category"),
                "amount": tx.get("amount"),
            }
            for tx in sorted(included, key=lambda t: t.get("date", ""), reverse=True)
        ],
        "excluded_rental_count": excluded_rental_count,
        "ledger_path": ledger_path,
        "plaid_env": plaid_env,
    }

    # ── Adriana rental file processing ───────────────────────────────────
    if drive_set:
        try:
            from adriana_parser import list_unprocessed_adriana_files, parse_adriana_file
            drive_svc = get_drive_service()
            wb_check = load_workbook(ledger_path)
            unprocessed = list_unprocessed_adriana_files(drive_svc, wb_check)
            wb_check.close()
            for fm in unprocessed:
                try:
                    a_txns = parse_adriana_file(drive_svc, fm)
                    a_result = write_spending_ledger(ledger_path, a_txns)
                    ym_key = f"{fm['year']}-{fm['month']:02d}"
                    set_meta_flag(ledger_path, f"adriana_processed:{ym_key}")
                    month_label = datetime(fm["year"], fm["month"], 1).strftime("%B %Y")
                    print(f"📋 Adriana {month_label}: {a_result['added']} added, {a_result['skipped']} skipped")
                except Exception as e_fm:
                    print(f"⚠️  Adriana '{fm['name']}': {e_fm} — skipping")
        except Exception as e_adriana:
            print(f"⚠️  Adriana sync failed (non-fatal): {e_adriana}")

    need_snapshot    = False
    snapshot_attachment: list[dict] | None = None
    year_month = date.today().strftime("%Y-%m")

    if is_cloud and file_id:
        upload_ok = False
        try:
            upload_ledger(file_id, ledger_path)
            upload_ok = True
        except Exception as e:
            print(f"⚠️  Drive upload failed: {e}")

        if upload_ok:
            try:
                wb_meta = load_workbook(ledger_path, read_only=True)
                last = get_last_snapshot_month(wb_meta)
                wb_meta.close()
                if last != year_month:
                    need_snapshot = True
                    with open(ledger_path, "rb") as _f:
                        encoded = base64.b64encode(_f.read()).decode()
                    snapshot_attachment = [{
                        "filename": f"cashflow-tracker-snapshot-{year_month}.xlsx",
                        "content": encoded,
                    }]
                    print(f"📸 New month ({year_month}) — ledger will be attached to sync email.")
                else:
                    print(f"📸 Snapshot already sent for {year_month} — no attachment this run.")
            except Exception as e:
                print(f"⚠️  _Meta read failed (non-fatal): {e}")

    email_ok = False
    try:
        send_sync_summary(summary, attachments=snapshot_attachment)
        email_ok = True
    except Exception as e:
        print(f"⚠️  Email failed: {e}")

    if need_snapshot and email_ok and is_cloud and file_id:
        try:
            wb_meta = load_workbook(ledger_path)
            set_last_snapshot_month(wb_meta, year_month)
            wb_meta.save(ledger_path)
            upload_ledger(file_id, ledger_path)
            print(f"📸 _Meta updated and re-uploaded for {year_month}")
        except Exception as e:
            print(f"⚠️  _Meta update/re-upload failed (non-fatal): {e}")

    excl_parts = [
        f"{len(txns)} from {mask} ({PRIMARY_EXCLUDED_MASKS[mask]})"
        for mask, txns in excluded_by_account.items() if txns
    ]
    if excl_parts:
        print(f"Excluded — {', '.join(excl_parts)}")
    print(f"✅ Sync complete — {added} added, {skipped} skipped, ${total_spend:,.2f} total spend")
    return summary


def _account_label(account: dict) -> str:
    subtype = account.get("subtype", "").lower()
    if "checking" in subtype:
        return "Checking"
    if "credit" in subtype:
        return "Credit Card"
    return account.get("name", "Unknown")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Spending Tracker sync")
    parser.add_argument(
        "--from-date",
        help="Start date for transaction fetch (YYYY-MM-DD). Defaults to 7 days ago.",
    )
    args = parser.parse_args()

    from_date = None
    if args.from_date:
        try:
            from_date = date.fromisoformat(args.from_date)
        except ValueError:
            parser.error("--from-date must be YYYY-MM-DD")
        if from_date > date.today():
            parser.error("--from-date cannot be in the future")
        if (date.today() - from_date).days > 730:
            print("⚠️  Warning: requesting more than 2 years of history — Plaid may limit results.")

    run_sync(from_date)
