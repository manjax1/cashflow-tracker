#!/usr/bin/env python3
"""Remove ONLY the excess duplicate transactions relative to a known-good
reference snapshot (e.g. a pre-Plaid-re-auth backup) — without deleting genuine
same-day repeats (subway taps, per-property utility bills) that legitimately
share a (date, description, account, amount, type) signature.

Plaid re-authentication re-issues NEW transaction_ids for the same real
transactions; the id-based dedup then re-adds them. But some signatures repeat
legitimately, so we can't just keep one per signature. Instead, for each
signature we keep exactly as many copies as the REFERENCE had, and remove the
extra ones (preferring to drop rows whose SourceRef isn't in the reference — the
re-auth copies). Signatures not in the reference (genuinely new since then) are
left alone except for exact-SourceRef duplicates, and flagged for your review.

DRY-RUN by default; backs up before writing, then rebuilds the summary sheets.

Usage:
    ls -t cashflow-tracker_BACKUP_*.xlsx | head          # pick a pre-re-auth backup
    python scripts/dedup_ledger.py --reference <backup.xlsx>
    python scripts/dedup_ledger.py --reference <backup.xlsx> --apply
"""
import os
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl
from dotenv import load_dotenv

load_dotenv()

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LEDGER = os.environ.get("SPENDING_LEDGER_FILE_PATH", os.path.join(ROOT, "cashflow-tracker.xlsx"))


def _sig(get):
    return (str(get("Date"))[:10], str(get("Description")).strip().lower(),
            str(get("Account")), round(abs(float(get("Amount") or 0)), 2), str(get("Type")))


def _read(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb["Transactions"].iter_rows(values_only=True))
    wb.close()
    hdr = [str(c) for c in rows[0]]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        if r[ix["Date"]] is None:
            continue
        get = lambda name, r=r: r[ix[name]]
        out.append((_sig(get), str(r[ix["SourceRef"]])))
    return out


def main():
    args = sys.argv[1:]
    if "--reference" not in args:
        sys.exit("Provide --reference <pre-re-auth backup .xlsx>.\n"
                 "   List candidates: ls -t cashflow-tracker_BACKUP_*.xlsx | head")
    ref_path = args[args.index("--reference") + 1]
    dry = "--apply" not in args

    ref = _read(ref_path)
    ref_counts = Counter(s for s, _ in ref)
    ref_refs = {sr for _, sr in ref}

    wb = openpyxl.load_workbook(LEDGER)
    ws = wb["Transactions"]
    hdr = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(hdr)}
    split_parents = set()
    if "Splits" in wb.sheetnames:
        sh = wb["Splits"]
        pref = [c.value for c in sh[1]].index("ParentRef")
        for r in sh.iter_rows(min_row=2, values_only=True):
            if r and r[pref]:
                split_parents.add(str(r[pref]))

    def val(ridx, name):
        return ws.cell(row=ridx, column=ix[name] + 1).value

    bysig = defaultdict(list)
    for ridx in range(2, ws.max_row + 1):
        if val(ridx, "Date") is None:
            continue
        bysig[_sig(lambda n, r=ridx: val(r, n))].append(ridx)

    remove, flagged = [], []
    inc_rm = exp_rm = 0.0
    for sig, idxs in bysig.items():
        R = ref_counts.get(sig, 0)
        if R > 0:
            if len(idxs) > R:
                # keep the reference copies (SourceRef in ref, or split parents) first
                ordered = sorted(idxs, key=lambda r: (str(val(r, "SourceRef")) in ref_refs
                                                      or str(val(r, "SourceRef")) in split_parents),
                                 reverse=True)
                for ridx in ordered[R:]:
                    remove.append(ridx)
        else:
            seen = set()
            for ridx in idxs:
                s = str(val(ridx, "SourceRef"))
                if s in seen:
                    remove.append(ridx)      # exact same id twice — always safe to drop
                else:
                    seen.add(s)
            if len(seen) > 1:
                flagged.append((sig, len(seen)))

    for ridx in remove:
        a = abs(float(val(ridx, "Amount") or 0))
        if str(val(ridx, "Type")) == "Income":
            inc_rm += a
        else:
            exp_rm += a

    print(f"Ledger:    {LEDGER}")
    print(f"Reference: {os.path.basename(ref_path)}  ({sum(ref_counts.values())} rows)")
    print(f"\nExcess duplicate rows to remove: {len(remove)}")
    print(f"  Inflated income removed:  ${inc_rm:,.2f}")
    print(f"  Inflated expense removed: ${exp_rm:,.2f}")
    if flagged:
        print(f"\n⚠ {len(flagged)} NEW-since-reference signature(s) repeat with different SourceRefs — "
              f"NOT auto-removed. Confirm whether these are real repeats or dupes:")
        for sig, n in sorted(flagged, key=lambda x: -x[0][3])[:15]:
            print(f"     {sig[0]}  ${sig[3]:>10,.2f} ×{n}  {sig[1][:38]}")

    if dry:
        print("\nDRY RUN — nothing changed. Re-run with --apply once the review list looks right.")
        return

    backup = LEDGER.replace(".xlsx", f"_BACKUP_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx")
    shutil.copy2(LEDGER, backup)
    for ridx in sorted(remove, reverse=True):
        ws.delete_rows(ridx, 1)
    wb.save(LEDGER)
    print(f"\nRemoved {len(remove)} rows. Backup: {os.path.basename(backup)}")
    sys.path.insert(0, os.path.join(ROOT, "src"))
    from ledger_writer import write_spending_ledger
    write_spending_ledger(LEDGER, [])
    print("Summary sheets rebuilt. Next: 'push' from the agent CLI (guarded).")


if __name__ == "__main__":
    main()
