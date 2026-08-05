"""Costco receipt ingestion — warehouse baskets, gas, and returns.

Costco warehouse charges are usually ALREADY in the ledger (BofA card is
Plaid-connected), so this module SPLITS the existing lump transaction into
item categories rather than appending. Receipts with no ledger match
(Citi card, or pre-ledger) are appended like Amazon.

Deterministic layer parses date/total/card/type; the LLM decodes Costco's
cryptic item abbreviations (KS = Kirkland Signature, etc.) and classifies.

Usage:
    python -m src.agent.costco extract --dry-run   # parse + classify receipts
    python -m src.agent.costco extract             # save to costco_data/
"""

import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta

from dotenv import load_dotenv

load_dotenv()

from . import ledger                 # noqa: E402
from . import invoices               # noqa: E402
from .tools import audit             # noqa: E402

INBOX = os.path.join(ledger.REPO_ROOT, "Costco-Purchases")
OUTDIR = os.path.join(ledger.REPO_ROOT, "costco_data")

COSTCO_TOOL = {
    "name": "record_receipt",
    "description": "Record a parsed Costco receipt's line items with categories.",
    "input_schema": {
        "type": "object",
        "properties": {
            "items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "decoded readable name"},
                        "raw": {"type": "string", "description": "original receipt abbreviation"},
                        "net_price": {"type": "number", "description": "extended price minus any instant-savings discount for this item"},
                        "category": {"type": "string"},
                        "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
                    },
                    "required": ["name", "net_price", "category", "confidence"],
                },
            },
        },
        "required": ["items"],
    },
}

COSTCO_SYSTEM = """You decode Costco warehouse receipts and classify each item.
Costco abbreviations: 'KS'/'KrkS' = Kirkland Signature; 'ORG' = organic;
'GRK' = Greek; produce/food names are heavily abbreviated (e.g. 'CHERRY TOV' =
cherry tomatoes, 'GRRY TRL' = a trail-mix, 'JW DBL BLACK' = Johnnie Walker
Double Black whisky). Decode to a readable name.

Line format: optional 'E', item number, NAME, extended price, then N or Y
(tax flag). A following 'coupon# / item# amount-' line is an instant-savings
DISCOUNT for that item — subtract it to get net_price. 'N @ price' lines above
an item indicate quantity (extended price already reflects it).

Classification rules (use ONLY the provided taxonomy names):
- Food, beverages, produce, pantry -> Groceries
- Alcohol (wine, beer, spirits — e.g. Johnnie Walker, La Crema, Sparrow Cabernet) -> Entertainment
- Costco Shop Card / gift card purchases -> 'Credit Card Payment' (stored value, not spend)
- Prescription eyewear / lenses / frames -> Health and Fitness
- Skincare, vitamins, supplements, hygiene, grooming -> Personal Care
- Apparel/shoes -> Clothing
- Cleaning/household supplies, batteries, hardware -> the best existing category
  (Groceries for consumable household; else Shopping)
- Fuel -> Transportation
Report net_price exactly (extended minus discount). Do not compute the total."""

# Vision tool — used for photographed/scanned receipts (no text layer) or
# unrecognized layouts, where the model must read the header fields too.
COSTCO_VISION_TOOL = {
    "name": "record_receipt_full",
    "description": "Record a Costco receipt read from an image: header fields plus line items.",
    "input_schema": {
        "type": "object",
        "properties": {
            "date": {"type": "string", "description": "purchase date, YYYY-MM-DD"},
            "total": {"type": "number", "description": "grand total; NEGATIVE if the receipt is a refund/return"},
            "card_last4": {"type": "string", "description": "last 4 digits of the tender card, if shown"},
            "type": {"type": "string", "enum": ["warehouse", "gas", "return"]},
            "instant_savings": {"type": "number", "description": "sum of discounts, else 0"},
            "items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "decoded readable name"},
                        "raw": {"type": "string"},
                        "net_price": {"type": "number", "description": "extended price minus any instant-savings; negative for a return line"},
                        "category": {"type": "string"},
                        "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
                    },
                    "required": ["name", "net_price", "category"],
                },
            },
        },
        "required": ["date", "total", "type", "items"],
    },
}


def parse_receipt_meta(text):
    """Deterministic: date, total, tender, card last-4, type.
    Handles two Costco layouts: warehouse/return receipts and gas receipts."""
    is_gas = bool(re.search(r"\b(pump|gallons|kirkland\s+signature\s+fuel|regular unlead|"
                            r"premium unlead)\b", text, re.I) or re.search(r"Total Sale", text))
    # date: warehouse "MM/DD/YYYY HH:MM"  or  gas "Date: MM/DD/YY"
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})\s+\d{2}:\d{2}", text)
    if m:
        date = f"{m.group(3)}-{m.group(1)}-{m.group(2)}"
    else:
        m = re.search(r"Date:\s*(\d{2})/(\d{2})/(\d{2})", text)
        date = f"20{m.group(3)}-{m.group(1)}-{m.group(2)}" if m else None
    # total: warehouse "**** TOTAL n"  or  gas "Total Sale $n"
    tot = re.search(r"\*+\s*TOTAL\s+([\d.,]+)(-?)", text)
    if tot:
        total = float(tot.group(1).replace(",", "")) * (-1 if tot.group(2) else 1)
    else:
        g = re.search(r"Total Sale\s+\$?([\d.,]+)", text)
        total = float(g.group(1).replace(",", "")) if g else None
    card = re.search(r"[X*]{5,}(\d{4})", text)
    is_return = bool(re.search(r"APPROVED\s*-\s*REFUND|TOTAL\s+[\d.,]+-", text, re.I))
    savings = re.search(r"INSTANT SAVINGS\s+\$?([\d.,]+)", text)
    return {
        "date": date, "total": total,
        "card_last4": card.group(1) if card else None,
        "type": "gas" if is_gas else ("return" if is_return else "warehouse"),
        "instant_savings": float(savings.group(1).replace(",", "")) if savings else 0,
    }


def _allocate(receipt):
    """Spread the receipt total across item net_prices (folds tax/rounding in),
    writing allocated_amount on each item. No-op if total or items are missing."""
    items = receipt.get("items", [])
    base = sum(abs(i["net_price"]) for i in items)
    total = abs(receipt["total"] if receipt.get("total") is not None else base)
    if base > 0:
        for i in items:
            i["allocated_amount"] = round(abs(i["net_price"]) * total / base, 2)
        drift = round(total - sum(i["allocated_amount"] for i in items), 2)
        if drift:
            big = max(items, key=lambda x: x["allocated_amount"])
            big["allocated_amount"] = round(big["allocated_amount"] + drift, 2)


IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic", ".heif")
_RAW_IMAGE_MEDIA = {".png": "image/png", ".webp": "image/webp", ".gif": "image/gif",
                    ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}


def _vision_source_block(path):
    """Build the Anthropic content block for a receipt file. PDFs go as a
    'document' block; image files as an 'image' block. HEIC/HEIF (and large
    photos) are normalized to JPEG and downscaled so the API accepts them."""
    import base64
    ext = os.path.splitext(path)[1].lower()
    with open(path, "rb") as f:
        raw = f.read()
    if ext == ".pdf":
        return {"type": "document",
                "source": {"type": "base64", "media_type": "application/pdf",
                           "data": base64.standard_b64encode(raw).decode()}}
    # Image: convert/downscale via Pillow (HEIC needs pillow-heif). Fall back to
    # raw bytes for common formats if Pillow isn't available.
    try:
        import io
        from PIL import Image
        if ext in (".heic", ".heif"):
            import pillow_heif
            pillow_heif.register_heif_opener()
        img = Image.open(io.BytesIO(raw)).convert("RGB")
        img.thumbnail((1600, 1600))                       # cap long edge; cheaper + reliable
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=88)
        data = base64.standard_b64encode(buf.getvalue()).decode()
        return {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": data}}
    except Exception:
        media = _RAW_IMAGE_MEDIA.get(ext)
        if not media:
            raise ValueError(f"{ext} images need Pillow/pillow-heif on the server; "
                             "please upload a JPEG, PNG, or PDF instead")
        return {"type": "image",
                "source": {"type": "base64", "media_type": media,
                           "data": base64.standard_b64encode(raw).decode()}}


def _extract_via_vision(path, client):
    """Read a photographed/scanned receipt (no text layer) or an unrecognized
    layout directly from the receipt image/PDF. Returns the full receipt dict."""
    source_block = _vision_source_block(path)
    cats = invoices.taxonomy()
    prompt = (
        "This is a photographed/scanned Costco receipt or store audit print-out. "
        "Read it carefully and call record_receipt_full with the header fields and "
        "EVERY line item. Notes: 'total' is the grand total (make it NEGATIVE only if "
        "the whole receipt is a refund); a line ending in '-' (e.g. '6.00-') is a return/"
        "credit, so its net_price is negative; decode Costco abbreviations to readable names.\n\n"
        "Taxonomy categories (use EXACT names):\n" + "\n".join(f"- {c}" for c in cats))
    resp = client.messages.create(
        model=invoices.MODEL, max_tokens=4000, system=COSTCO_SYSTEM,
        tools=[COSTCO_VISION_TOOL], tool_choice={"type": "tool", "name": "record_receipt_full"},
        messages=[{"role": "user", "content": [source_block, {"type": "text", "text": prompt}]}])
    v = next(b.input for b in resp.content if b.type == "tool_use")
    return {
        "source_file": os.path.basename(path),
        "date": v.get("date"),
        "total": v.get("total"),
        "card_last4": v.get("card_last4"),
        "type": v.get("type") or "warehouse",
        "instant_savings": v.get("instant_savings") or 0,
        "items": v.get("items", []),
        "_extracted_at": datetime.now().isoformat(timespec="seconds"),
        "_via": "vision",
    }


def extract_receipt(path, client):
    # A bare image file (phone photo) has no text layer to parse — read it with
    # the vision model directly.
    if os.path.splitext(path)[1].lower() in IMAGE_EXTS:
        receipt = _extract_via_vision(path, client)
        _allocate(receipt)
        return receipt
    text = invoices.read_pdf(path)
    meta = parse_receipt_meta(text)
    # Scanned image-only PDF or unrecognized layout: the text layer is empty and
    # the regex can't find the total -> read the receipt image with the vision model.
    if meta.get("total") is None or len((text or "").strip()) < 40:
        receipt = _extract_via_vision(path, client)
        _allocate(receipt)
        return receipt

    receipt = {"source_file": os.path.basename(path), **meta,
               "_extracted_at": datetime.now().isoformat(timespec="seconds")}
    if meta["type"] == "gas":
        receipt["items"] = [{"name": "Costco gas", "net_price": abs(meta["total"] or 0),
                             "category": "Transportation", "confidence": "high"}]
        return receipt
    cats = invoices.taxonomy()
    prompt = ("Taxonomy categories (use EXACT names):\n"
              + "\n".join(f"- {c}" for c in cats)
              + f"\n\nCostco receipt text:\n---\n{text[:8000]}\n---\n"
              "Extract and classify every line item with record_receipt.")
    resp = client.messages.create(
        model=invoices.MODEL, max_tokens=4000, system=COSTCO_SYSTEM,
        tools=[COSTCO_TOOL], tool_choice={"type": "tool", "name": "record_receipt"},
        messages=[{"role": "user", "content": prompt}])
    receipt["items"] = next(b.input for b in resp.content if b.type == "tool_use")["items"]
    _allocate(receipt)   # reconcile: scale item net_prices to the receipt total
    return receipt


def receipt_id(meta):
    return f"costco-{meta['date']}-{abs(meta['total'] or 0):.2f}".replace(".", "")


def extract_all(dry_run=False):
    import anthropic
    if not os.path.isdir(INBOX):
        sys.exit(f"No Costco folder at {INBOX}")
    os.makedirs(OUTDIR, exist_ok=True)
    client = anthropic.Anthropic()
    files = sorted(f for f in os.listdir(INBOX) if f.endswith(".pdf"))
    counts = {"warehouse": 0, "gas": 0, "return": 0}
    for fname in files:
        meta = parse_receipt_meta(invoices.read_pdf(os.path.join(INBOX, fname)))
        rid = receipt_id(meta)
        if os.path.exists(os.path.join(OUTDIR, rid + ".json")):
            continue
        print(f"  ⚙ {fname}  ({meta['type']}, {meta['date']}, ${meta['total']}, card ...{meta['card_last4']})")
        r = extract_receipt(os.path.join(INBOX, fname), client)
        r["receipt_id"] = rid
        counts[meta["type"]] += 1
        for i in r.get("items", []):
            c = {"high": "", "medium": " (?)", "low": " (??)"}[i["confidence"]]
            print(f"      ${i.get('allocated_amount', i['net_price']):>8.2f}  {i['category']}{c}  {i['name'][:44]}")
        if not dry_run:
            with open(os.path.join(OUTDIR, rid + ".json"), "w") as f:
                json.dump(r, f, indent=2)
    audit("costco_extracted", counts)
    print(f"\n{sum(counts.values())} receipts: {counts}"
          + (" (dry run — nothing saved)" if dry_run else f" → {OUTDIR}"))


def _match_charge(cc, date, amount, ttype):
    if amount is None or not date:
        return []
    return [t for t in cc if abs(t["Amount"] - abs(amount)) < 0.01
            and t["Type"] == ttype and t["Date"][:7] == date[:7]
            and abs(int(t["Date"][8:10]) - int(date[8:10])) <= 4]


def split_one(receipt, ledger_path):
    """Match one extracted receipt to a discrete ledger charge and write its
    splits. Returns a status dict. Used by the web upload endpoint.
    Idempotent: re-splitting an already-split charge is refused."""
    import openpyxl
    ledger.LEDGER_PATH = ledger_path
    ledger._cache["mtime"] = None
    ledger._splits_cache["mtime"] = None
    txns = ledger.load_transactions()
    ledger_start = min(t["Date"] for t in txns)
    cc = [t for t in txns if "costco" in t["Description"].lower()
          and t["Account"] == "Credit Card"]
    d, tot, typ = receipt.get("date"), receipt.get("total"), receipt.get("type")
    if tot is None:
        return {"status": "no_total",
                "reason": "couldn't read the total from this receipt image — try a sharper, "
                          "flatter photo, or the emailed/printed PDF"}
    if not d:
        return {"status": "no_date", "reason": "couldn't read the purchase date from this receipt"}
    ttype = "Income" if (tot or 0) < 0 else "Expense"
    hits = _match_charge(cc, d, tot, ttype)
    if not hits:
        return {"status": "no_match",
                "reason": (f"no matching Costco charge of ${abs(tot):,.2f} on/near {d} in the "
                           "ledger — it may not have synced from the bank yet (try after the "
                           "next sync), or this card isn't one of the connected accounts"
                           if d >= ledger_start else "purchase predates the ledger")}
    ref = str(hits[0]["SourceRef"])
    if ref in ledger.load_splits():
        return {"status": "already_split", "parent_ref": ref}
    if typ == "gas" and hits[0]["Category"] == "Transportation":
        return {"status": "gas_no_split", "note": "gas already categorized as Transportation"}
    items = receipt["items"]
    base = sum(i.get("allocated_amount", 0) for i in items)
    drift = round(abs(tot) - base, 2)
    if abs(drift) > 0.02 and items:
        big = max(items, key=lambda x: x.get("allocated_amount", 0))
        big["allocated_amount"] = round(big.get("allocated_amount", 0) + drift, 2)
    wb = openpyxl.load_workbook(ledger_path)
    if "Splits" in wb.sheetnames:
        ws = wb["Splits"]
    else:
        ws = wb.create_sheet("Splits")
        ws.append(["ParentRef", "Item", "Category", "Amount", "ReceiptID", "CreatedAt"])
    now = datetime.now().isoformat(timespec="seconds")
    breakdown = defaultdict(float)
    for i in items:
        amt = round(i.get("allocated_amount", 0), 2)
        ws.append([ref, i["name"][:60], i["category"], amt, receipt["receipt_id"], now])
        breakdown[i["category"]] += amt
    wb.save(ledger_path)
    ledger._cache["mtime"] = None
    ledger._splits_cache["mtime"] = None
    audit("costco_receipt_split", {"receipt": receipt["receipt_id"], "parent_ref": ref,
                                   "items": len(items)})
    return {"status": "split", "parent_ref": ref, "date": hits[0]["Date"],
            "charge": hits[0]["Amount"], "items": len(items),
            "breakdown": {k: round(v, 2) for k, v in breakdown.items()}}


# ── Deferred reconciliation: queue receipts whose charge hasn't posted yet ──
_PENDING_HEADERS = ["ReceiptID", "Date", "Total", "Type", "Items", "QueuedAt", "ReceiptJSON"]


def queue_pending(receipt, ledger_path):
    """Save an already-extracted receipt that has no matching charge yet, into a
    'PendingReceipts' sheet in the ledger. The daily sync retries it once the
    charge posts. Keyed by receipt_id — re-queuing the same receipt replaces its
    row (idempotent). The full classified receipt is stored, so reconciliation
    later is pure matching (no re-reading the image)."""
    import openpyxl
    rid = receipt.get("receipt_id") or receipt_id(receipt)
    wb = openpyxl.load_workbook(ledger_path)
    if "PendingReceipts" in wb.sheetnames:
        ws = wb["PendingReceipts"]
    else:
        ws = wb.create_sheet("PendingReceipts")
        ws.append(_PENDING_HEADERS)
    for r in range(ws.max_row, 1, -1):                 # drop any prior row for this receipt
        if str(ws.cell(row=r, column=1).value) == rid:
            ws.delete_rows(r, 1)
    ws.append([rid, receipt.get("date"), receipt.get("total"), receipt.get("type"),
               len(receipt.get("items", [])),
               datetime.now().isoformat(timespec="seconds"), json.dumps(receipt)])
    wb.save(ledger_path)
    ledger._cache["mtime"] = None
    audit("costco_receipt_queued", {"receipt": rid, "total": receipt.get("total")})
    return {"status": "queued", "receipt_id": rid, "date": receipt.get("date"),
            "total": receipt.get("total"),
            "reason": "saved — it'll auto-split when the charge posts (usually within a day or two)"}


def list_pending(ledger_path):
    """Return the queued receipts awaiting a charge (for the web drawer)."""
    import openpyxl
    wb = openpyxl.load_workbook(ledger_path, read_only=True)
    if "PendingReceipts" not in wb.sheetnames:
        wb.close()
        return []
    out = []
    for r in wb["PendingReceipts"].iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            out.append({"receipt_id": str(r[0]), "date": r[1], "total": r[2],
                        "type": r[3], "items": r[4], "queued_at": r[5]})
    wb.close()
    return out


def remove_pending(ledger_path, receipt_ids):
    """Delete queued receipts by id (manual clear from the web). Returns count."""
    import openpyxl
    ids = {str(x) for x in receipt_ids}
    wb = openpyxl.load_workbook(ledger_path)
    if "PendingReceipts" not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb["PendingReceipts"]
    removed = 0
    for rr in range(ws.max_row, 1, -1):
        if str(ws.cell(row=rr, column=1).value) in ids:
            ws.delete_rows(rr, 1)
            removed += 1
    if removed:
        wb.save(ledger_path)
        ledger._cache["mtime"] = None
        audit("costco_pending_cleared", {"count": removed, "ids": sorted(ids)})
    return removed


PENDING_EXPIRY_DAYS = int(os.getenv("COSTCO_PENDING_EXPIRY_DAYS", "100"))


def reconcile_pending(ledger_path):
    """Retry every queued receipt against the current ledger charges. Resolved
    ones (split / already split / gas) are removed. Receipts that still haven't
    matched after PENDING_EXPIRY_DAYS (default 100) are expired and removed so
    receipts on un-connected cards don't linger forever. Called by the daily
    sync. Returns a summary dict."""
    import openpyxl
    wb = openpyxl.load_workbook(ledger_path, read_only=True)
    if "PendingReceipts" not in wb.sheetnames:
        wb.close()
        return {"processed": 0, "split": 0, "resolved": 0, "expired": 0,
                "still_pending": 0, "details": []}
    rows = [r for r in wb["PendingReceipts"].iter_rows(min_row=2, values_only=True) if r and r[0]]
    wb.close()
    ji = _PENDING_HEADERS.index("ReceiptJSON")
    qi = _PENDING_HEADERS.index("QueuedAt")
    cutoff = datetime.now() - timedelta(days=PENDING_EXPIRY_DAYS)

    resolved, expired, details, split_n = [], [], [], 0
    for r in rows:
        try:
            receipt = json.loads(r[ji])
        except (TypeError, ValueError, json.JSONDecodeError):
            continue
        receipt.setdefault("receipt_id", str(r[0]))
        res = split_one(receipt, ledger_path)
        d = {"receipt_id": str(r[0]), "status": res["status"]}
        if res["status"] == "split":
            d.update({"date": res.get("date"), "charge": res.get("charge"),
                      "items": res.get("items"), "breakdown": res.get("breakdown", {})})
        if res["status"] in ("split", "already_split", "gas_no_split"):
            resolved.append(str(r[0]))
            split_n += res["status"] == "split"
        else:
            # Unmatched — expire it if it's been queued longer than the limit.
            try:
                too_old = datetime.fromisoformat(str(r[qi])) < cutoff
            except (TypeError, ValueError):
                too_old = False
            if too_old:
                expired.append(str(r[0]))
                d.update({"status": "expired", "date": receipt.get("date"),
                          "total": receipt.get("total")})
        details.append(d)

    to_remove = set(resolved) | set(expired)
    if to_remove:
        wb2 = openpyxl.load_workbook(ledger_path)
        ws2 = wb2["PendingReceipts"]
        for rr in range(ws2.max_row, 1, -1):
            if str(ws2.cell(row=rr, column=1).value) in to_remove:
                ws2.delete_rows(rr, 1)
        wb2.save(ledger_path)
        ledger._cache["mtime"] = None
        ledger._splits_cache["mtime"] = None
        audit("costco_pending_reconciled",
              {"split": split_n, "resolved": len(resolved), "expired": len(expired),
               "still_pending": len(rows) - len(to_remove)})
    return {"processed": len(rows), "split": split_n, "resolved": len(resolved),
            "expired": len(expired), "still_pending": len(rows) - len(to_remove),
            "details": details}


def reconcile(apply=False):
    """Match saved Costco receipts to discrete ledger charges and split them
    into item categories via the Splits sheet. Receipts with no discrete
    charge (Citi aggregate, or pre-ledger) are reported and left untouched."""
    import shutil
    import openpyxl

    txns = ledger.load_transactions()
    ledger_start = min(t["Date"] for t in txns)
    cc = [t for t in txns if "costco" in t["Description"].lower()
          and t["Account"] == "Credit Card"]
    existing_splits = ledger.load_splits()

    to_write, skipped = [], {"aggregate": 0, "prewindow": 0, "already": 0, "gas_ok": 0}
    for f in sorted(os.listdir(OUTDIR)):
        if not f.endswith(".json"):
            continue
        r = json.load(open(os.path.join(OUTDIR, f)))
        d, tot, typ = r["date"], r["total"], r["type"]
        ttype = "Income" if (tot or 0) < 0 else "Expense"
        hits = _match_charge(cc, d, tot, ttype)
        if not hits:
            skipped["prewindow" if d < ledger_start else "aggregate"] += 1
            continue
        ref = str(hits[0]["SourceRef"])
        if ref in existing_splits:
            skipped["already"] += 1
            continue
        # gas already sits in Transportation as a single line — no split needed
        if typ == "gas" and hits[0]["Category"] == "Transportation":
            skipped["gas_ok"] += 1
            continue
        items = r["items"]
        base = sum(i.get("allocated_amount", 0) for i in items)
        drift = round(abs(tot) - base, 2)
        if abs(drift) > 0.02 and items:
            big = max(items, key=lambda x: x.get("allocated_amount", 0))
            big["allocated_amount"] = round(big.get("allocated_amount", 0) + drift, 2)
        for i in items:
            to_write.append((ref, i["name"][:60], i["category"],
                             round(i.get("allocated_amount", 0), 2), r["receipt_id"]))

    from collections import defaultdict
    bycat = defaultdict(float)
    for _, _, cat, amt, _ in to_write:
        bycat[cat] += amt
    print(f"{len(to_write)} split rows across {len(set(w[0] for w in to_write))} charges")
    print(f"skipped: {skipped}")
    print("\nSplit spend by category (moving OUT of the lump charges):")
    for c, v in sorted(bycat.items(), key=lambda kv: -kv[1]):
        print(f"  {c:<26} ${v:>9.2f}")
    if not apply:
        print("\n(dry run — nothing written; add 'apply' to write the Splits sheet)")
        return
    if input(f"\nWrite {len(to_write)} split rows to the Splits sheet? [y/N] ").strip().lower() != "y":
        print("Cancelled.")
        return
    backup = ledger.LEDGER_PATH.replace(".xlsx", f"_BACKUP_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx")
    shutil.copy2(ledger.LEDGER_PATH, backup)
    wb = openpyxl.load_workbook(ledger.LEDGER_PATH)
    if "Splits" in wb.sheetnames:
        ws = wb["Splits"]
    else:
        ws = wb.create_sheet("Splits")
        ws.append(["ParentRef", "Item", "Category", "Amount", "ReceiptID", "CreatedAt"])
    now = datetime.now().isoformat(timespec="seconds")
    for ref, item, cat, amt, rid in to_write:
        ws.append([ref, item, cat, amt, rid, now])
    wb.save(ledger.LEDGER_PATH)
    ledger._cache["mtime"] = None
    audit("costco_splits_written", {"rows": len(to_write), "backup": os.path.basename(backup)})
    print(f"✅ {len(to_write)} split rows written. Backup: {os.path.basename(backup)}")
    print("Next: 'push' from the agent CLI to sync Drive.")


if __name__ == "__main__":
    args = sys.argv[1:]
    if args and args[0] == "extract":
        extract_all(dry_run="--dry-run" in args)
    elif args and args[0] == "reconcile":
        reconcile(apply="apply" in args)
    else:
        print(__doc__)
