import os
import sys
import threading
from datetime import date, datetime, timedelta
from flask import Flask, jsonify, request

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils import clean_env, resolve_ledger_path

app = Flask(__name__)

_sync_lock = threading.Lock()
_sync_running = False
_last_run: dict = {}


def _get_local_time() -> str:
    tz_str = clean_env(os.getenv("TZ", "America/Los_Angeles"), "TZ")
    try:
        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo(tz_str))
    except Exception:
        now = datetime.utcnow()
    return now.isoformat()


def _wants_html() -> bool:
    best = request.accept_mimetypes.best_match(["application/json", "text/html"])
    return best == "text/html"


def _html_response(status_code: int, data: dict) -> tuple:
    if status_code == 409:
        body = f"<p>⏳ {data.get('message', 'A sync is already in progress, please wait')}</p>"
    elif status_code >= 400:
        body = f"<p>❌ {data.get('message', 'Error')}</p>"
    else:
        from_d = data.get("from_date") or "default (7 days ago)"
        to_d = data.get("to_date", "today")
        body = f"<p>✅ Sync started for {from_d} to {to_d}</p>"
        note = data.get("note")
        if note:
            body += f"<p><em>Note: {note}</em></p>"
    html = f"<!doctype html><html><body>{body}</body></html>"
    return html, status_code, {"Content-Type": "text/html"}


def _respond(status_code: int, data: dict):
    if _wants_html():
        return _html_response(status_code, data)
    return jsonify(data), status_code


def _parse_date_param(value: str, name: str):
    """Returns (date_obj, error_response) — exactly one will be non-None."""
    try:
        return date.fromisoformat(value), None
    except ValueError:
        err = {"status": "error", "message": f"Invalid {name} '{value}' — expected YYYY-MM-DD"}
        return None, _respond(400, err)


@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "service": "spending-tracker",
        "environment": clean_env(os.getenv("PLAID_ENV", "production"), "PLAID_ENV"),
        "timestamp": _get_local_time(),
        "sync_running": _sync_running,
        "last_run": _last_run,
    })


@app.route("/sync", methods=["POST"])
def sync():
    global _sync_running, _last_run

    if not _sync_lock.acquire(blocking=False):
        return jsonify({"status": "busy", "message": "A sync is already running"}), 409

    body = request.get_json(silent=True) or {}
    from_date_str = body.get("from_date")
    to_date_str = body.get("to_date")

    def _run():
        global _sync_running, _last_run
        _sync_running = True
        try:
            from main import run_sync
            from_date = date.fromisoformat(from_date_str) if from_date_str else None
            to_date = date.fromisoformat(to_date_str) if to_date_str else None
            result = run_sync(from_date, to_date)
            _last_run = result
        except Exception as e:
            _last_run = {"error": str(e), "timestamp": _get_local_time()}
            print(f"Sync error: {e}")
        finally:
            _sync_running = False
            _sync_lock.release()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    return jsonify({"status": "started"})


@app.route("/sync/run", methods=["GET"])
def sync_run():
    global _sync_running, _last_run

    today = date.today()

    from_date_str = request.args.get("from_date")
    to_date_str = request.args.get("to_date")

    from_date = None
    if from_date_str:
        from_date, err = _parse_date_param(from_date_str, "from_date")
        if err:
            return err

    to_date = today
    to_date_clamped = False
    if to_date_str:
        to_date, err = _parse_date_param(to_date_str, "to_date")
        if err:
            return err
        if to_date > today:
            to_date = today
            to_date_clamped = True

    if from_date and to_date < from_date:
        return _respond(400, {
            "status": "error",
            "message": f"to_date ({to_date}) must not be before from_date ({from_date})",
        })

    if not _sync_lock.acquire(blocking=False):
        return _respond(409, {"status": "busy", "message": "A sync is already in progress, please wait"})

    # Capture resolved dates for the thread closure
    resolved_from = from_date
    resolved_to = to_date

    def _run():
        global _sync_running, _last_run
        _sync_running = True
        try:
            from main import run_sync
            result = run_sync(resolved_from, resolved_to)
            _last_run = result
        except Exception as e:
            _last_run = {"error": str(e), "timestamp": _get_local_time()}
            print(f"Sync error: {e}")
        finally:
            _sync_running = False
            _sync_lock.release()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()

    resp = {
        "status": "started",
        "from_date": str(from_date) if from_date else None,
        "to_date": str(to_date),
    }
    if to_date_clamped:
        resp["note"] = f"to_date was in the future and has been clamped to today ({today})"
    return _respond(200, resp)


@app.route("/sync/test", methods=["POST"])
def sync_test():
    results = {}

    # Test Plaid
    try:
        from plaid_client import PlaidClient
        client = PlaidClient()
        access_token = clean_env(os.getenv("PLAID_ACCESS_TOKEN"), "PLAID_ACCESS_TOKEN")
        if access_token and client.verify_access_token(access_token):
            results["plaid"] = "connected"
        else:
            results["plaid"] = "error: invalid or missing access token"
    except Exception as e:
        results["plaid"] = f"error: {e}"

    # Test Drive
    try:
        from drive_sync import get_drive_service
        svc = get_drive_service()
        svc.files().list(pageSize=1, fields="files(id)").execute()
        results["drive"] = "connected"
    except Exception as e:
        results["drive"] = f"error: {e}"

    all_ok = all(v == "connected" for v in results.values())
    results["status"] = "ok" if all_ok else "degraded"
    return jsonify(results)


@app.route("/ledger/info")
def ledger_info():
    """Diagnostic: read the live ledger file and return real row-level totals.
    Completely read-only — opens with read_only=True, data_only=True, no writes.
    """
    ledger_path, _ = resolve_ledger_path()
    if not ledger_path:
        return jsonify({"status": "error", "message": "SPENDING_LEDGER_FILE_PATH not set"}), 500
    if not os.path.exists(ledger_path):
        return jsonify({"status": "error", "message": f"Ledger not found at {ledger_path}"}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(ledger_path, read_only=True, data_only=True)
        ws = wb["Transactions"]

        # Map column positions from the actual header row so renames don't silently break this.
        # Expected order: Date, Description, Account, Category, Type, Amount, IncludeInNet, SourceRef
        raw_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {name: i for i, name in enumerate(raw_header) if name}
        date_idx = col.get("Date",         0)
        type_idx = col.get("Type",         4)
        amt_idx  = col.get("Amount",       5)
        net_idx  = col.get("IncludeInNet", 6)

        total   = 0
        income  = 0.0
        expense = 0.0
        earliest = None
        latest   = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:          # skip blank rows
                continue
            total += 1

            # Parse date — openpyxl may return datetime, date, or ISO string
            dv = row[date_idx]
            try:
                if hasattr(dv, "date"):
                    d = dv.date()
                elif isinstance(dv, str):
                    from datetime import date as _date
                    d = _date.fromisoformat(str(dv)[:10])
                else:
                    d = None
                if d:
                    if earliest is None or d < earliest:
                        earliest = d
                    if latest is None or d > latest:
                        latest = d
            except Exception:
                pass

            # Sum Income / Expense (skip rows excluded from net)
            amt     = row[amt_idx]
            include = row[net_idx]
            if amt is not None and include is not False:
                ttype = str(row[type_idx] or "")
                if ttype == "Income":
                    income  += float(amt)
                elif ttype == "Expense":
                    expense += float(amt)

        wb.close()

        return jsonify({
            "status":       "ok",
            "ledger_path":  ledger_path,
            "row_count":    total,
            "earliest_date": str(earliest) if earliest else None,
            "latest_date":   str(latest)   if latest   else None,
            "income":  round(income,           2),
            "expense": round(expense,          2),
            "net":     round(income - expense, 2),
            "generated_at": _get_local_time(),
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/transactions")
def transactions():
    """Read-only: return ledger transactions filtered by date range, category, and account."""
    today = date.today()

    since_str     = request.args.get("since")
    from_date_str = request.args.get("from_date")
    to_date_str   = request.args.get("to_date")

    if since_str and (from_date_str or to_date_str):
        return jsonify({"status": "error",
                        "message": "'since' cannot be combined with from_date or to_date"}), 400

    if since_str:
        from_date, err = _parse_date_param(since_str, "since")
        if err:
            return err
        to_date = today
    else:
        if from_date_str:
            from_date, err = _parse_date_param(from_date_str, "from_date")
            if err:
                return err
        else:
            from_date = today - timedelta(days=30)

        if to_date_str:
            to_date, err = _parse_date_param(to_date_str, "to_date")
            if err:
                return err
        else:
            to_date = today

    if to_date < from_date:
        return jsonify({"status": "error",
                        "message": f"to_date ({to_date}) must not be before from_date ({from_date})"}), 400

    category_filter = request.args.get("category")
    account_filter  = request.args.get("account")

    ledger_path, _ = resolve_ledger_path()
    if not ledger_path:
        return jsonify({"status": "error", "message": "SPENDING_LEDGER_FILE_PATH not set"}), 500
    if not os.path.exists(ledger_path):
        return jsonify({"status": "error", "message": f"Ledger not found at {ledger_path}"}), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(ledger_path, read_only=True, data_only=True)
        ws = wb["Transactions"]

        raw_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col      = {name: i for i, name in enumerate(raw_header) if name}
        date_idx = col.get("Date",         0)
        desc_idx = col.get("Description",  1)
        acct_idx = col.get("Account",      2)
        cat_idx  = col.get("Category",     3)
        type_idx = col.get("Type",         4)
        amt_idx  = col.get("Amount",       5)
        net_idx  = col.get("IncludeInNet", 6)

        txns          = []
        total_income  = 0.0
        total_expense = 0.0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            dv = row[date_idx]
            try:
                d = dv.date() if hasattr(dv, "date") else date.fromisoformat(str(dv)[:10])
            except Exception:
                continue

            if d < from_date or d > to_date:
                continue

            cat   = str(row[cat_idx]  or "")
            acct  = str(row[acct_idx] or "")

            if category_filter and cat  != category_filter:
                continue
            if account_filter  and acct != account_filter:
                continue

            ttype = str(row[type_idx] or "")
            amt   = float(row[amt_idx] or 0)

            txns.append({
                "date":        str(d),
                "description": str(row[desc_idx] or ""),
                "account":     acct,
                "category":    cat,
                "type":        ttype,
                "amount":      round(amt, 2),
            })

            if row[net_idx] is not False:
                if ttype == "Income":
                    total_income  += amt
                elif ttype == "Expense":
                    total_expense += amt

        wb.close()

        txns.sort(key=lambda t: t["date"], reverse=True)

        active_filters = {k: v for k, v in
                          [("category", category_filter), ("account", account_filter)] if v}

        return jsonify({
            "from_date": str(from_date),
            "to_date":   str(to_date),
            "filters":   active_filters,
            "summary": {
                "count":         len(txns),
                "total_income":  round(total_income,              2),
                "total_expense": round(total_expense,             2),
                "net":           round(total_income - total_expense, 2),
            },
            "transactions": txns,
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", 8080)))
