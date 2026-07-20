"""Deterministic data-access layer over the cashflow ledger (xlsx).

Design principle: ALL arithmetic happens here. The model interprets;
it never computes. Every function returns plain JSON-serializable data.
"""

import json
import os
import statistics
from collections import defaultdict
from datetime import date, datetime, timedelta

import openpyxl
from dotenv import load_dotenv

load_dotenv()

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LEDGER_PATH = os.environ.get(
    "SPENDING_LEDGER_FILE_PATH",
    os.path.join(REPO_ROOT, "cashflow-tracker.xlsx"),
)
RULES_PATH = os.path.join(REPO_ROOT, "spending_rules.json")

_cache = {"mtime": None, "rows": None}
_splits_cache = {"mtime": None, "splits": None}


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def load_transactions():
    """Load and cache the Transactions sheet as a list of dicts."""
    mtime = os.path.getmtime(LEDGER_PATH)
    if _cache["mtime"] == mtime:
        return _cache["rows"]
    wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
    ws = wb["Transactions"]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(rows_iter)]
    rows = []
    for r in rows_iter:
        if r[0] is None:
            continue
        row = dict(zip(header, r))
        row["Date"] = _parse_date(row["Date"]).isoformat()
        row["Amount"] = float(row["Amount"] or 0)
        row["IncludeInNet"] = bool(row.get("IncludeInNet", True))
        rows.append(row)
    wb.close()
    _cache.update(mtime=mtime, rows=rows)
    return rows


def load_splits():
    """Read the optional 'Splits' sheet: parent SourceRef -> [item rows].
    Non-destructive overlay; absent sheet just means no splits."""
    mtime = os.path.getmtime(LEDGER_PATH)
    if _splits_cache["mtime"] == mtime and _splits_cache["splits"] is not None:
        return _splits_cache["splits"]
    splits = defaultdict(list)
    wb = openpyxl.load_workbook(LEDGER_PATH, read_only=True, data_only=True)
    if "Splits" in wb.sheetnames:
        rows = wb["Splits"].iter_rows(values_only=True)
        header = [str(h) for h in next(rows)]
        for r in rows:
            if r[0] is None:
                continue
            d = dict(zip(header, r))
            splits[str(d["ParentRef"])].append({
                "item": d.get("Item", ""), "category": d["Category"],
                "amount": float(d["Amount"] or 0)})
    wb.close()
    _splits_cache.update(mtime=mtime, splits=splits)
    return splits


def effective_rows():
    """Transactions with split-parents expanded into their item rows.
    A charge that has splits contributes its per-category item amounts instead
    of its single original category. Everything downstream aggregates on this."""
    rows = load_transactions()
    splits = load_splits()
    if not splits:
        return rows
    out = []
    for t in rows:
        sp = splits.get(str(t["SourceRef"]))
        if not sp:
            out.append(t)
            continue
        for s in sp:
            child = dict(t)
            child["Category"] = s["category"]
            child["Amount"] = s["amount"]
            child["Description"] = f"{t['Description'][:40]} :: {s['item']}"
            child["_split_of"] = t["SourceRef"]
            out.append(child)
    return out


def _in_net(t):
    return t["IncludeInNet"]


def _filter(rows, start_date=None, end_date=None, category=None, account=None,
            tx_type=None, min_amount=None, max_amount=None, search=None):
    out = []
    for t in rows:
        if start_date and t["Date"] < start_date:
            continue
        if end_date and t["Date"] > end_date:
            continue
        if category and category.lower() not in t["Category"].lower():
            continue
        if account and t["Account"].lower() != account.lower():
            continue
        if tx_type and t["Type"].lower() != tx_type.lower():
            continue
        if min_amount is not None and t["Amount"] < min_amount:
            continue
        if max_amount is not None and t["Amount"] > max_amount:
            continue
        if search and search.lower() not in t["Description"].lower():
            continue
        out.append(t)
    return out


# ----------------------------- read tools -----------------------------

def query_transactions(start_date, end_date, category=None, account=None,
                       tx_type=None, min_amount=None, max_amount=None,
                       search=None, limit=100):
    rows = _filter(effective_rows(), start_date, end_date, category,
                   account, tx_type, min_amount, max_amount, search)
    rows = sorted(rows, key=lambda t: t["Date"], reverse=True)
    total = len(rows)
    inc = sum(t["Amount"] for t in rows if t["Type"] == "Income" and t["IncludeInNet"])
    exp = sum(t["Amount"] for t in rows if t["Type"] == "Expense" and t["IncludeInNet"])
    return {
        "count": total,
        "truncated": total > limit,
        "totals": {"income": round(inc, 2), "expense": round(exp, 2),
                   "net": round(inc - exp, 2)},
        "transactions": rows[:limit],
    }


def get_cashflow_summary(start_date, end_date, group_by="category", category=None):
    """Aggregates income/expense/net. group_by: category|month|account|type."""
    rows = _filter(effective_rows(), start_date, end_date, category=category)
    rows = [t for t in rows if _in_net(t)]

    def key(t):
        if group_by == "month":
            return t["Date"][:7]
        if group_by == "account":
            return t["Account"]
        if group_by == "type":
            return t["Type"]
        return t["Category"]

    groups = defaultdict(lambda: {"income": 0.0, "expense": 0.0, "count": 0})
    totals = {"income": 0.0, "expense": 0.0}
    for t in rows:
        g = groups[key(t)]
        side = "income" if t["Type"] == "Income" else "expense"
        g[side] += t["Amount"]
        g["count"] += 1
        totals[side] += t["Amount"]

    out_groups = []
    for k, v in groups.items():
        out_groups.append({
            "group": k,
            "income": round(v["income"], 2),
            "expense": round(v["expense"], 2),
            "net": round(v["income"] - v["expense"], 2),
            "count": v["count"],
        })
    out_groups.sort(key=lambda g: g["expense"] + g["income"], reverse=True)

    rental_exp = sum(g["expense"] for g in out_groups if g["group"].startswith("Rental"))
    rental_inc = sum(g["income"] for g in out_groups if g["group"].startswith("Rental"))
    return {
        "start_date": start_date,
        "end_date": end_date,
        "group_by": group_by,
        "total_income": round(totals["income"], 2),
        "total_expense": round(totals["expense"], 2),
        "net": round(totals["income"] - totals["expense"], 2),
        "rental_rollup": {"income": round(rental_inc, 2), "expense": round(rental_exp, 2),
                          "net": round(rental_inc - rental_exp, 2)},
        "groups": out_groups,
    }


def get_trends(metric="net", granularity="month", lookback_periods=6, category=None):
    """Deterministic time series with period-over-period deltas + trailing avg."""
    rows = [t for t in effective_rows() if _in_net(t)]
    if category:
        rows = [t for t in rows if category.lower() in t["Category"].lower()]

    def period_of(t):
        if granularity == "quarter":
            y, m = t["Date"][:4], int(t["Date"][5:7])
            return f"{y}-Q{(m - 1) // 3 + 1}"
        return t["Date"][:7]

    per = defaultdict(lambda: {"income": 0.0, "expense": 0.0})
    for t in rows:
        side = "income" if t["Type"] == "Income" else "expense"
        per[period_of(t)][side] += t["Amount"]

    periods = sorted(per.keys())[-lookback_periods:]
    series, prev_val = [], None
    for p in periods:
        v = per[p]
        val = {"income": v["income"], "expenses": v["expense"],
               "net": v["income"] - v["expense"]}[metric if metric != "expense" else "expenses"]
        val = round(val, 2)
        delta = round(val - prev_val, 2) if prev_val is not None else None
        pct = round(100 * delta / abs(prev_val), 1) if delta is not None and prev_val else None
        series.append({"period": p, "value": val, "delta": delta, "delta_pct": pct})
        prev_val = val
    # Mark the newest period as partial if ledger coverage ends mid-period,
    # so charts can drop it and the model won't misread a mid-month dip.
    if series:
        import calendar
        all_dates = sorted(t["Date"] for t in load_transactions())
        cov_end = _parse_date(all_dates[-1])
        if granularity == "quarter":
            q_end_month = ((cov_end.month - 1) // 3 + 1) * 3
            period_end = date(cov_end.year, q_end_month,
                              calendar.monthrange(cov_end.year, q_end_month)[1])
        else:
            period_end = date(cov_end.year, cov_end.month,
                              calendar.monthrange(cov_end.year, cov_end.month)[1])
        last_period_of_cov = (f"{cov_end.year}-Q{(cov_end.month - 1) // 3 + 1}"
                              if granularity == "quarter" else all_dates[-1][:7])
        if series[-1]["period"] == last_period_of_cov and cov_end < period_end:
            series[-1]["partial"] = True

    complete = [s for s in series if not s.get("partial")]
    vals = [s["value"] for s in complete]
    return {
        "metric": metric, "granularity": granularity, "category": category,
        "series": series,
        "trailing_average": round(sum(vals) / len(vals), 2) if vals else 0,
        "min": min(vals) if vals else 0, "max": max(vals) if vals else 0,
        "note": "trailing_average/min/max exclude the partial current period",
    }


def list_categories():
    rows = effective_rows()
    cats = defaultdict(lambda: {"count": 0, "income": 0.0, "expense": 0.0,
                                "first_seen": "9999", "last_seen": "0000"})
    for t in rows:
        c = cats[t["Category"]]
        c["count"] += 1
        side = "income" if t["Type"] == "Income" else "expense"
        c[side] += t["Amount"]
        c["first_seen"] = min(c["first_seen"], t["Date"])
        c["last_seen"] = max(c["last_seen"], t["Date"])
    out = [{"category": k, "count": v["count"], "income": round(v["income"], 2),
            "expense": round(v["expense"], 2), "first_seen": v["first_seen"],
            "last_seen": v["last_seen"]} for k, v in cats.items()]
    out.sort(key=lambda c: c["count"], reverse=True)
    return {"categories": out, "note": "Categories prefixed 'Rental - ' roll up into rental subtotals."}


def get_category_rules(keyword_filter=None):
    with open(RULES_PATH) as f:
        rules = json.load(f)
    if keyword_filter:
        kf = keyword_filter.lower()
        rules = [r for r in rules if kf in r["keyword"].lower()
                 or kf in r["category"].lower() or kf in r.get("note", "").lower()]
    return {"count": len(rules), "rules": rules[:200],
            "matching": "case-insensitive, longest keyword wins; fallback to Plaid PFC, then 'Uncategorized'"}


def find_anomalies(lookback_days=30):
    """Deterministic checks. The model explains/prioritizes; it does not detect."""
    rows = load_transactions()
    if not rows:
        return {"findings": []}
    end = max(t["Date"] for t in rows)
    start = (_parse_date(end) - timedelta(days=lookback_days)).isoformat()
    window = [t for t in rows if start <= t["Date"] <= end]
    findings = []

    # 1. Possible duplicates (same date, amount, description)
    seen = defaultdict(list)
    for t in window:
        seen[(t["Date"], t["Amount"], t["Description"])].append(t)
    for k, v in seen.items():
        if len(v) > 1 and v[0]["Amount"] >= 10:  # small repeats (transit taps etc.) are usually legit
            findings.append({"type": "possible_duplicate", "severity": "high",
                             "detail": f"{len(v)}x '{k[2][:60]}' on {k[0]} for ${k[1]:.2f}",
                             "source_refs": [t["SourceRef"] for t in v]})

    # 2. Unusually large vs category median (>=3x, category needs >=5 samples)
    by_cat = defaultdict(list)
    for t in rows:
        if t["Type"] == "Expense" and t["Amount"] > 0:
            by_cat[t["Category"]].append(t["Amount"])
    for t in window:
        amts = by_cat.get(t["Category"], [])
        if t["Type"] == "Expense" and len(amts) >= 5:
            med = statistics.median(amts)
            if med > 0 and t["Amount"] >= 3 * med:
                findings.append({"type": "large_transaction", "severity": "medium",
                                 "detail": f"'{t['Description'][:60]}' ${t['Amount']:.2f} on {t['Date']} "
                                           f"is {t['Amount']/med:.1f}x the {t['Category']} median (${med:.2f})",
                                 "source_refs": [t["SourceRef"]]})

    # 3. Missing rental income this month vs prior months
    rental_inc_months = {t["Date"][:7] for t in rows
                         if t["Category"] == "Rental - Income"}
    cur_month = end[:7]
    if rental_inc_months and cur_month not in rental_inc_months:
        findings.append({"type": "missing_expected_income", "severity": "high",
                         "detail": f"No 'Rental - Income' transactions recorded in {cur_month}; "
                                   f"present in {len(rental_inc_months)} prior months",
                         "source_refs": []})

    # 4. Category spend spike: window total >= 2x trailing 3-month avg
    hist_start = (_parse_date(start) - timedelta(days=90)).isoformat()
    cur = defaultdict(float)
    hist = defaultdict(float)
    for t in rows:
        if t["Type"] != "Expense" or not _in_net(t):
            continue
        if start <= t["Date"] <= end:
            cur[t["Category"]] += t["Amount"]
        elif hist_start <= t["Date"] < start:
            hist[t["Category"]] += t["Amount"]
    for cat, amt in cur.items():
        avg = hist.get(cat, 0) / 3
        if avg > 50 and amt >= 2 * avg:
            findings.append({"type": "category_spike", "severity": "medium",
                             "detail": f"{cat}: ${amt:.2f} in last {lookback_days}d vs "
                                       f"${avg:.2f}/mo trailing average (>=2x)",
                             "source_refs": []})

    # 5. Uncategorized backlog
    uncat = [t for t in window if "Uncategorized" in t["Category"]]
    if len(uncat) >= 5:
        findings.append({"type": "uncategorized_backlog", "severity": "low",
                         "detail": f"{len(uncat)} uncategorized transactions in window "
                                   f"totaling ${sum(t['Amount'] for t in uncat):.2f}",
                         "source_refs": [t["SourceRef"] for t in uncat[:10]]})

    sev_rank = {"high": 0, "medium": 1, "low": 2}
    findings.sort(key=lambda f: sev_rank[f["severity"]])
    return {"window": {"start": start, "end": end}, "finding_count": len(findings),
            "findings": findings}
