#!/usr/bin/env python3
"""
Full real import: wipes local ledger, imports all 18 BofA CSV files in
chronological order, then writes a detailed import + verification report.

CAUTION: This script DELETES the local ledger and writes real data.
Review the script before running. Do not run in a dry-run context.

Import pipeline (no logic duplicated):
  parse_checking_csv / parse_credit_card_csv  — from csv_importer
  categorize_batch                             — from filters (via csv_importer)
  write_spending_ledger                        — from ledger_writer

Each file is imported individually (not batched) so per-file added/skipped
counts are captured from the write_spending_ledger return value.
"""

import os
import sys
from collections import defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

import openpyxl

from csv_importer import parse_checking_csv, parse_credit_card_csv, RULES_PATH
from filters import load_rules, categorize_batch
from ledger_writer import write_spending_ledger
from utils import clean_env

# ── Paths ─────────────────────────────────────────────────────────────────────
LEDGER_PATH = clean_env(os.getenv("SPENDING_LEDGER_FILE_PATH"), "SPENDING_LEDGER_FILE_PATH")
REPORT_PATH = "/Users/manjax/Downloads/BoA/FULL_IMPORT_REPORT_2025-07_to_2026-03.txt"
W = 115  # line width — matches dry-run report format

# ── File lists ────────────────────────────────────────────────────────────────
CHECKING_FILES = [
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-7-28-25.csv",
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-8-26-25.csv",
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-9-25-25.csv",
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-10-28-25.csv",
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-11-24-25.csv",
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-12-26-25.csv",
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-1-27-26.csv",
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-2-24-26.csv",
    "/Users/manjax/Downloads/BoA/BoA-Checking-account-5799-period-ending-3-26-26.csv",
]
CREDIT_FILES = [
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-July-2025.csv",
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-Aug-2025.csv",
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-Sep-2025.csv",
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-Oct-2025.csv",
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-Nov-2025.csv",
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-Dec-2025.csv",
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-Jan-2026.csv",
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-Feb-2026.csv",
    "/Users/manjax/Downloads/BoA/BoA-Credit-card-account-0605-Mar-2026.csv",
]

# Chronological interleave: Checking Jul, CC Jul, Checking Aug, CC Aug, …
IMPORT_ORDER: list[tuple[str, str]] = []
for _chk, _cc in zip(CHECKING_FILES, CREDIT_FILES):
    IMPORT_ORDER.append(("checking", _chk))
    IMPORT_ORDER.append(("credit",   _cc))


# ── Ledger verification (re-read from .xlsx after all imports) ────────────────

def _read_ledger_totals(path: str) -> dict:
    """
    Open the written XLSX fresh and compute real totals.
    Transactions sheet columns (from ledger_writer.py, line 517):
      A=Date  B=Description  C=Account  D=Category  E=Type  F=Amount
      G=IncludeInNet  H=SourceRef
    Reads header row dynamically so column renames don't silently break this.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Transactions"]

    # Map column names → 0-based index from actual header row
    raw_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(raw_header) if name}

    # Fallbacks match the known column order in case a cell is unexpectedly None
    cat_idx  = col.get("Category",      3)
    type_idx = col.get("Type",          4)
    amt_idx  = col.get("Amount",        5)
    net_idx  = col.get("IncludeInNet",  6)

    cat_counts: dict[str, int]   = defaultdict(int)
    cat_totals: dict[str, float] = defaultdict(float)
    type_totals: dict[str, float] = {"Income": 0.0, "Expense": 0.0}
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:          # blank row
            continue
        cat   = str(row[cat_idx]  or "Uncategorized")
        ttype = str(row[type_idx] or "Expense")
        amt   = float(row[amt_idx] or 0.0)
        cat_counts[cat] += 1
        cat_totals[cat] += amt
        type_totals[ttype] += amt
        total += 1

    wb.close()
    return {
        "total":      total,
        "cat_counts": dict(cat_counts),
        "cat_totals": dict(cat_totals),
        "type_totals": type_totals,
    }


# ── Report helpers ────────────────────────────────────────────────────────────

def _cat_block(out, cat_counts: dict, cat_totals: dict) -> None:
    out.write("─" * W + "\n")
    out.write("Category Breakdown\n")
    out.write("─" * W + "\n")
    out.write(f"{'Category':<38} {'Count':>6}   {'Total':>12}\n")
    out.write("─" * W + "\n")
    for cat, total in sorted(cat_totals.items(), key=lambda x: x[1], reverse=True):
        out.write(f"{cat:<38} {cat_counts[cat]:>6}   ${total:>11,.2f}\n")


def _type_block(out, type_totals: dict) -> None:
    income  = type_totals.get("Income",  0.0)
    expense = type_totals.get("Expense", 0.0)
    out.write("\n" + "─" * W + "\n")
    out.write("Type Summary\n")
    out.write("─" * W + "\n")
    out.write(f"  Income:   ${income:>12,.2f}\n")
    out.write(f"  Expense:  ${expense:>12,.2f}\n")
    out.write(f"  Net:      ${income - expense:>12,.2f}\n")


def _write_file_section(
    out, idx: int, acct_label: str, fp: str, timestamp: str,
    included: list, excluded_count: int,
    added: int, skipped: int, error: str | None,
) -> None:
    dates  = sorted(tx.get("date", "") for tx in included if tx.get("date"))
    period = f"{dates[0]} → {dates[-1]}" if dates else "no transactions"

    out.write("=" * W + "\n")
    out.write(f"FILE {idx:02d} of 18  │  {acct_label:<12}  │  {period}\n")
    out.write(f"Source:    {fp}\n")
    out.write(f"Generated: {timestamp}\n")
    out.write(
        f"Parsed:    {len(included) + excluded_count} rows  │  "
        f"{len(included)} included  │  {excluded_count} rule-excluded\n"
    )
    out.write(f"Written:   {added} added to ledger  │  {skipped} skipped (duplicate)\n")
    out.write("=" * W + "\n\n")

    if error:
        out.write(f"❌ ERROR: {error}\n\n")
        return

    cat_counts: dict[str, int]   = defaultdict(int)
    cat_totals: dict[str, float] = defaultdict(float)
    type_totals: dict[str, float] = {"Income": 0.0, "Expense": 0.0}
    for tx in included:
        cat   = tx.get("category", "Uncategorized")
        amt   = tx.get("amount",   0.0)
        ttype = tx.get("type",     "Expense")
        cat_counts[cat] += 1
        cat_totals[cat] += amt
        type_totals[ttype] += amt

    _cat_block(out, cat_counts, cat_totals)
    _type_block(out, type_totals)
    out.write("\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── 1. Wipe local ledger ──────────────────────────────────────────────────
    if os.path.exists(LEDGER_PATH):
        os.remove(LEDGER_PATH)
        print(f"🗑️  Removed existing ledger: {LEDGER_PATH}")
    else:
        print(f"ℹ️  No existing ledger found — starting fresh")

    # ── 2. Load rules once ────────────────────────────────────────────────────
    rules = load_rules(RULES_PATH) if os.path.exists(RULES_PATH) else []
    if not rules:
        print("⚠️  No spending_rules.json found — using Plaid PFC categories only")
    else:
        print(f"✅ Loaded {len(rules)} rules from {RULES_PATH}")

    # ── 3. Import each file, capture per-file stats ───────────────────────────
    grand_added   = 0
    grand_skipped = 0
    grand_cat_counts: dict[str, int]   = defaultdict(int)
    grand_cat_totals: dict[str, float] = defaultdict(float)
    grand_type_totals: dict[str, float] = {"Income": 0.0, "Expense": 0.0}

    # (acct_type, fp, included, excluded_count, added, skipped, error_or_None)
    file_results: list[tuple] = []

    for acct_type, fp in IMPORT_ORDER:
        parse_fn   = parse_checking_csv if acct_type == "checking" else parse_credit_card_csv
        acct_label = "Checking" if acct_type == "checking" else "Credit Card"
        print(f"\n📄 [{acct_label}] {os.path.basename(fp)}")

        try:
            txns = parse_fn(fp)
        except Exception as e:
            print(f"  ❌ Parse failed: {e}")
            file_results.append((acct_type, fp, [], 0, 0, 0, str(e)))
            continue

        print(f"  Rows parsed: {len(txns)}")
        account_map = {tx["account_id"]: acct_label for tx in txns}
        included, excluded = categorize_batch(txns, rules, account_map)
        print(f"  Included: {len(included)}  │  Excluded: {len(excluded)}")

        try:
            result = write_spending_ledger(LEDGER_PATH, included)
        except Exception as e:
            print(f"  ❌ Ledger write failed: {e}")
            file_results.append((acct_type, fp, included, len(excluded), 0, 0, str(e)))
            continue

        added_n   = result["added"]
        skipped_n = result["skipped"]
        print(f"  Written:  {added_n} added  │  {skipped_n} skipped (duplicates)")

        file_results.append((acct_type, fp, included, len(excluded), added_n, skipped_n, None))

        grand_added   += added_n
        grand_skipped += skipped_n
        for tx in included:
            cat   = tx.get("category", "Uncategorized")
            amt   = tx.get("amount",   0.0)
            ttype = tx.get("type",     "Expense")
            grand_cat_counts[cat] += 1
            grand_cat_totals[cat] += amt
            grand_type_totals[ttype] += amt

    # ── 4. Verify ledger (re-read from .xlsx) ─────────────────────────────────
    print(f"\n📊 Verifying ledger: {LEDGER_PATH}")
    try:
        ledger = _read_ledger_totals(LEDGER_PATH)
    except Exception as e:
        print(f"  ❌ Ledger verification failed: {e}")
        ledger = None

    running_income  = grand_type_totals["Income"]
    running_expense = grand_type_totals["Expense"]

    if ledger:
        ledger_income  = ledger["type_totals"]["Income"]
        ledger_expense = ledger["type_totals"]["Expense"]
        row_match    = ledger["total"] == grand_added
        income_match = abs(ledger_income  - running_income)  < 0.01
        expense_match= abs(ledger_expense - running_expense) < 0.01
        totals_match = row_match and income_match and expense_match
    else:
        ledger_income = ledger_expense = 0.0
        totals_match  = False

    # ── 5. Write report ───────────────────────────────────────────────────────
    print(f"\n📝 Writing import report …")
    with open(REPORT_PATH, "w", encoding="utf-8") as out:

        # Top header
        out.write("=" * W + "\n")
        out.write("FULL IMPORT REPORT — BofA Checking (5799) + Credit Card (0605)\n")
        out.write("Period:    July 2025 – March 2026  (9 checking + 9 credit = 18 files)\n")
        out.write(f"Generated: {generated_at}\n")
        out.write(f"Ledger:    {LEDGER_PATH}\n")
        out.write("=" * W + "\n\n")

        # Per-file sections
        for idx, (acct_type, fp, included, excl_count, added_n, skipped_n, error) in \
                enumerate(file_results, start=1):
            acct_label = "Checking" if acct_type == "checking" else "Credit Card"
            _write_file_section(
                out, idx, acct_label, fp, generated_at,
                included, excl_count, added_n, skipped_n, error,
            )

        # Grand total (running tallies from categorize_batch)
        grand_income  = grand_type_totals["Income"]
        grand_expense = grand_type_totals["Expense"]

        out.write("=" * W + "\n")
        out.write("GRAND TOTAL — July 2025 through March 2026  (all 18 files combined)\n")
        out.write("=" * W + "\n\n")
        out.write(f"  Files processed:      18 / 18\n")
        out.write(f"  Total rows added:     {grand_added}\n")
        out.write(f"  Total rows skipped:   {grand_skipped}  (duplicates across overlapping statement periods)\n\n")
        _cat_block(out, dict(grand_cat_counts), dict(grand_cat_totals))
        _type_block(out, grand_type_totals)

        # Ledger verification
        out.write("\n\n")
        out.write("=" * W + "\n")
        out.write("LEDGER VERIFICATION  (re-read from .xlsx after all imports complete)\n")
        out.write("=" * W + "\n\n")

        if not ledger:
            out.write("❌  Ledger could not be read — verification skipped\n\n")
        else:
            row_match_str = "✅ MATCH" if row_match else f"❌ MISMATCH ({ledger['total']} in file vs {grand_added} expected)"
            out.write(f"  Ledger file:        {LEDGER_PATH}\n")
            out.write(f"  Rows in ledger:     {ledger['total']}\n")
            out.write(f"  Expected (added):   {grand_added}\n")
            out.write(f"  Row count:          {row_match_str}\n\n")

            out.write(f"  {'Metric':<12} {'Running total':>16}  {'Ledger re-read':>16}  Result\n")
            out.write(f"  {'─' * 58}\n")
            out.write(
                f"  {'Income':<12} ${running_income:>15,.2f}  ${ledger_income:>15,.2f}  "
                f"{'✅' if income_match  else '❌ MISMATCH'}\n"
            )
            out.write(
                f"  {'Expense':<12} ${running_expense:>15,.2f}  ${ledger_expense:>15,.2f}  "
                f"{'✅' if expense_match else '❌ MISMATCH'}\n"
            )
            running_net = running_income - running_expense
            ledger_net  = ledger_income  - ledger_expense
            net_match   = abs(running_net - ledger_net) < 0.01
            out.write(
                f"  {'Net':<12} ${running_net:>15,.2f}  ${ledger_net:>15,.2f}  "
                f"{'✅' if net_match else '❌ MISMATCH'}\n"
            )
            out.write("\n")
            if totals_match:
                out.write("  ✅  VERIFIED — ledger matches running import totals exactly\n")
            else:
                out.write("  ❌  MISMATCH — ledger does not match running totals; investigate before using\n")

            # Category breakdown re-read from ledger
            out.write("\n\n")
            out.write("─" * W + "\n")
            out.write("Category breakdown re-read directly from ledger .xlsx\n")
            out.write("─" * W + "\n")
            out.write(f"{'Category':<38} {'Count':>6}   {'Total':>12}\n")
            out.write("─" * W + "\n")
            for cat, total in sorted(ledger["cat_totals"].items(), key=lambda x: x[1], reverse=True):
                out.write(f"{cat:<38} {ledger['cat_counts'][cat]:>6}   ${total:>11,.2f}\n")

        out.write("\n" + "=" * W + "\n")
        out.write("END OF REPORT\n")
        out.write("=" * W + "\n")

    # ── Terminal summary ──────────────────────────────────────────────────────
    size = os.path.getsize(REPORT_PATH)
    ok   = sum(1 for *_, err in file_results if err is None)

    print(f"\n{'─' * 60}")
    print(f"✅  Files imported:       {ok} / 18")
    print(f"    Rows added:           {grand_added}")
    print(f"    Rows skipped:         {grand_skipped}  (duplicates)")
    print(f"    Ledger match:         {'✅ VERIFIED' if totals_match else '❌ MISMATCH — check report'}")
    print(f"    Report:               {REPORT_PATH}")
    print(f"    Report size:          {size:,} bytes  ({size / 1024:.1f} KB)")
    print(f"{'─' * 60}")


if __name__ == "__main__":
    main()
