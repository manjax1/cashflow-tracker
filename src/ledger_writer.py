import os
from calendar import monthrange as _cal_mrange
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Arial", size=10)
SEC_FILL     = PatternFill("solid", fgColor="D6E4F0")
SEC_FONT     = Font(name="Arial", bold=True, size=10, color="1F3864")
SUB_FILL     = PatternFill("solid", fgColor="BDD7EE")
SUB_FONT     = Font(name="Arial", bold=True, size=10)
GRAND_FILL   = PatternFill("solid", fgColor="1F3864")
GRAND_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
NET_FILL     = PatternFill("solid", fgColor="E2EFDA")
NET_FONT     = Font(name="Arial", bold=True, size=10, color="375623")
NO_FILL          = PatternFill(fill_type=None)
HIGHLIGHT_FILL   = PatternFill("solid", fgColor="FFC107")  # amber — monthly category winner
HIGHLIGHT_FONT   = Font(name="Arial", bold=True, size=10)

CURRENCY_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_PALETTE = [
    "FFF2CC", "D9EAD3", "CFE2F3", "F4CCCC", "EAD1DC",
    "D9D2E9", "FCE5CD", "D0E4F1", "E6F4EA", "FFF3E0",
    "F3E5F5", "E8F5E9", "FFF8E1", "E3F2FD", "FCE4EC",
]


def _category_fill(category: str) -> PatternFill:
    idx = hash(category or "") % len(_PALETTE)
    return PatternFill("solid", fgColor=_PALETTE[idx])


def _header_row(ws, columns: list, row: int = 1):
    for col, title in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _tx_dedup_key(tx: dict) -> str:
    tid = tx.get("transaction_id")
    if tid:
        return str(tid)
    return f"{tx.get('date')}|{tx.get('name', '').strip().lower()}|{abs(tx.get('amount', 0)):.2f}"


def _existing_keys(ws) -> set:
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            # Column H (index 7) stores transaction_id for CSV-imported rows.
            source_ref = str(row[7]).strip() if len(row) > 7 and row[7] else ""
            if source_ref:
                keys.add(source_ref)
            date_val = str(row[0]) if row[0] else ""
            desc_val = str(row[1]).strip().lower() if row[1] else ""
            amt_val  = f"{abs(float(row[5])):.2f}" if row[5] is not None else "0.00"
            keys.add(f"{date_val}|{desc_val}|{amt_val}")
    return keys


def _clear_ws(ws, from_row: int = 2):
    for row in ws.iter_rows(min_row=from_row):
        for cell in row:
            cell.value = None
            cell.fill  = NO_FILL
            cell.font  = BODY_FONT


def _build_summary_sheet(wb, year: int):
    ws = wb.create_sheet("Monthly Summary")
    ws.column_dimensions["A"].width = 30
    ws["A1"].value = "Category"
    ws["A1"].font  = HEADER_FONT
    ws["A1"].fill  = HEADER_FILL
    total_col = len(MONTHS) + 2
    for m_idx, month in enumerate(MONTHS, 2):
        c = ws.cell(row=2, column=m_idx, value=month)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(m_idx)].width = 11
    tc = ws.cell(row=2, column=total_col, value="Total")
    tc.font = HEADER_FONT
    tc.fill = HEADER_FILL
    tc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(total_col)].width = 13
    avg_col = total_col + 1
    ac = ws.cell(row=2, column=avg_col, value="Trailing 12-Mo Avg")
    ac.font = HEADER_FONT
    ac.fill = HEADER_FILL
    ac.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(avg_col)].width = 16
    ws.freeze_panes = "B3"
    return ws


def _build_ytd_sheet(wb):
    ws = wb.create_sheet("YTD Summary")
    _header_row(ws, ["Category", "YTD Total", "% of Total"], row=1)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    return ws


def _build_yoy_sheet(wb):
    ws = wb.create_sheet("YoY Trends")
    ws["A1"] = "Year-over-year comparison requires 12+ months of data to populate Last Year column."
    ws["A1"].font = Font(name="Arial", italic=True, color="888888")
    _header_row(ws, ["Category", "This Year YTD", "Last Year Same Period", "$ Change", "% Change"], row=2)
    ws.column_dimensions["A"].width = 34
    for col in "BCDE":
        ws.column_dimensions[col].width = 18
    return ws


def _build_drilldown_sheet(wb):
    ws = wb.create_sheet("Drill-down")
    ws.column_dimensions["A"].width = 12   # Section 1: Date
    ws.column_dimensions["B"].width = 38   # Section 1: Description
    ws.column_dimensions["C"].width = 18   # Section 1: Account
    ws.column_dimensions["D"].width = 34   # Section 1: Category
    ws.column_dimensions["E"].width = 10   # Section 1: Type
    ws.column_dimensions["F"].width = 14   # Section 1: Amount
    ws.column_dimensions["H"].width = 12   # Section 2: Date
    ws.column_dimensions["I"].width = 38   # Section 2: Description
    ws.column_dimensions["J"].width = 18   # Section 2: Account
    ws.column_dimensions["K"].width = 34   # Section 2: Category
    ws.column_dimensions["L"].width = 10   # Section 2: Type
    ws.column_dimensions["M"].width = 14   # Section 2: Amount
    ws.column_dimensions["N"].width = 14   # helper: month list
    ws.column_dimensions["O"].width = 36   # helper: category list
    return ws


def _refresh_summary_formulas(wb, year: int):
    """Rebuild all summary sheets with Income / Rental Expense / Personal Expense sections."""
    tx_ws = wb["Transactions"]

    today      = date.today().strftime("%Y-%m-%d")
    ytd_start  = f"{year}-01-01"
    last_year  = year - 1
    ly_start   = f"{last_year}-01-01"
    try:
        ly_end = date(last_year, date.today().month, date.today().day).strftime("%Y-%m-%d")
    except ValueError:
        ly_end = f"{last_year}-12-31"

    # Scan Transactions for the actual (year, month) pairs present in the data.
    # This drives dynamic column generation so July–December 2025 data isn't
    # silently zeroed by formulas that only match the current calendar year.
    ym_set: set = set()
    for _scan_row in tx_ws.iter_rows(min_row=2, values_only=True):
        _dv = _scan_row[0]
        if not _dv:
            continue
        try:
            _d = _dv.date() if hasattr(_dv, "date") else date.fromisoformat(str(_dv))
            ym_set.add((_d.year, _d.month))
        except Exception:
            continue
    # [(2025, 7), (2025, 8), …, (2026, 3)] — chronological; fall back to current year if empty
    active_months: list[tuple[int, int]] = sorted(ym_set) or [(year, m) for m in range(1, 13)]
    total_col = len(active_months) + 2   # A=category, dynamic month cols, last col=Total
    avg_col   = total_col + 1            # Trailing 12-Mo Avg (rightmost column)
    # AVERAGE range spans the last min(N, 12) month columns (never the Total column).
    # Month cols run from 2 to total_col-1; trailing-12 start = max(2, total_col-12).
    _avg_start = get_column_letter(max(2, total_col - 12))
    _avg_end   = get_column_letter(total_col - 1)

    # Scan Transactions sheet: build ordered category lists by type.
    # Skip categories where IncludeInNet (col G) is False — those rows are
    # visible in Transactions but intentionally excluded from all summary totals.
    income_cats, rental_exp_cats, personal_exp_cats = [], [], []
    seen: set = set()
    for row in tx_ws.iter_rows(min_row=2, values_only=True):
        cat, tx_type, include_in_net = row[3], row[4], row[6]
        if not cat or cat in seen:
            continue
        if include_in_net is False:
            seen.add(cat)  # prevent re-processing but don't add to summary lists
            continue
        seen.add(cat)
        if tx_type == "Income":
            income_cats.append(cat)
        elif tx_type == "Expense":
            if cat.startswith("Rental - "):
                rental_exp_cats.append(cat)
            else:
                personal_exp_cats.append(cat)

    # ── helper: SUMPRODUCT formula builder ──────────────────────────────────
    # Column G (IncludeInNet) is a boolean; multiplying by it excludes FALSE rows
    # from all totals — this filters out "One-Off - Non-Recurring" rows automatically.
    def _sp(cat, tx_type, start, end):
        return (
            f'=SUMPRODUCT((Transactions!D$2:D$10000="{cat}")*'
            f'(Transactions!E$2:E$10000="{tx_type}")*'
            f'(Transactions!A$2:A$10000>="{start}")*'
            f'(Transactions!A$2:A$10000<="{end}")*'
            f'Transactions!G$2:G$10000*'
            f'Transactions!F$2:F$10000)'
        )

    def _month_end_str(yr: int, mn: int) -> str:
        return f"{yr}-{mn:02d}-{_cal_mrange(yr, mn)[1]:02d}"

    # ════════════════════════════════════════════════════════════════════════
    # MONTHLY SUMMARY
    # ════════════════════════════════════════════════════════════════════════
    ms  = wb["Monthly Summary"]
    _clear_ws(ms, from_row=2)   # also rebuilds the month-header row
    cur = 3                     # first data row; row 2 is rebuilt below as dynamic headers

    # Rebuild month header row (row 2) with "MMM YYYY" labels per active month
    month_labels = [f"{MONTHS[mn - 1]} {yr}" for yr, mn in active_months]
    _header_row(ms, ["Category"] + month_labels + ["Total", "Trailing 12-Mo Avg"], row=2)
    for col_idx in range(2, avg_col + 1):
        ms.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center")
        if col_idx == avg_col:
            ms.column_dimensions[get_column_letter(col_idx)].width = 16
        elif col_idx == total_col:
            ms.column_dimensions[get_column_letter(col_idx)].width = 13
        else:
            ms.column_dimensions[get_column_letter(col_idx)].width = 11
    ms.freeze_panes = "B3"

    # tracks row numbers for subtotals so we can reference them in grand totals
    ms_income_subtotal    = None
    ms_rental_subtotal    = None
    ms_personal_subtotal  = None
    # populated by _ms_cat_row; used later to apply per-month highlight fills
    ms_cat_row_map: dict = {}  # category string → Monthly Summary row number

    def _ms_sec_header(label):
        nonlocal cur
        for col in range(1, avg_col + 1):
            c = ms.cell(row=cur, column=col)
            c.fill  = SEC_FILL
            c.font  = SEC_FONT if col == 1 else Font(name="Arial", size=10)
            c.value = label if col == 1 else None
        cur += 1

    def _ms_cat_row(cat, tx_type):
        nonlocal cur
        r = cur
        ms_cat_row_map[cat] = r
        cf = _category_fill(cat)
        ms.cell(row=r, column=1, value=cat).fill = cf
        ms.cell(row=r, column=1).font = Font(name="Arial", size=10)
        for col_idx, (yr, mn) in enumerate(active_months, start=2):
            c = ms.cell(row=r, column=col_idx,
                        value=_sp(cat, tx_type, f"{yr}-{mn:02d}-01", _month_end_str(yr, mn)))
            c.number_format = CURRENCY_FMT
            c.font  = Font(name="Arial", size=10)
            c.fill  = cf
        b = get_column_letter(2)
        m = get_column_letter(1 + len(active_months))
        ms.cell(row=r, column=total_col,
                value=f"=SUM({b}{r}:{m}{r})").number_format = CURRENCY_FMT
        avg_c = ms.cell(row=r, column=avg_col,
                        value=f"=AVERAGE({_avg_start}{r}:{_avg_end}{r})")
        avg_c.number_format = CURRENCY_FMT
        avg_c.font  = Font(name="Arial", size=10)
        avg_c.fill  = cf
        cur += 1
        return r

    def _ms_subtotal(label, data_rows, fill=SUB_FILL, font=SUB_FONT):
        nonlocal cur
        r = cur
        ms.cell(row=r, column=1, value=label).font = font
        ms.cell(row=r, column=1).fill = fill
        for col in range(2, total_col + 1):
            ltr  = get_column_letter(col)
            refs = "+".join(f"{ltr}{dr}" for dr in data_rows)
            c = ms.cell(row=r, column=col, value=f"={refs}")
            c.number_format = CURRENCY_FMT
            c.fill = fill
            c.font = font
        avg_c = ms.cell(row=r, column=avg_col,
                        value=f"=AVERAGE({_avg_start}{r}:{_avg_end}{r})")
        avg_c.number_format = CURRENCY_FMT
        avg_c.fill = fill
        avg_c.font = font
        cur += 1
        return r

    # Income section
    ms_income_rows = []
    if income_cats:
        _ms_sec_header("── INCOME ──")
        for cat in income_cats:
            ms_income_rows.append(_ms_cat_row(cat, "Income"))
        ms_income_subtotal = _ms_subtotal("TOTAL INCOME", ms_income_rows)
        cur += 1

    # Rental Expense section
    ms_rental_rows = []
    if rental_exp_cats:
        _ms_sec_header("── RENTAL EXPENSES ──")
        for cat in rental_exp_cats:
            ms_rental_rows.append(_ms_cat_row(cat, "Expense"))
        ms_rental_subtotal = _ms_subtotal("TOTAL RENTAL EXPENSE", ms_rental_rows)
        cur += 1

    # Personal Expense section
    ms_personal_rows = []
    if personal_exp_cats:
        _ms_sec_header("── PERSONAL EXPENSES ──")
        for cat in personal_exp_cats:
            ms_personal_rows.append(_ms_cat_row(cat, "Expense"))
        ms_personal_subtotal = _ms_subtotal("TOTAL PERSONAL EXPENSE", ms_personal_rows)
        cur += 1

    # Grand Total Expense
    exp_subs = [r for r in [ms_rental_subtotal, ms_personal_subtotal] if r]
    ms_grand_expense = None
    if exp_subs:
        ms_grand_expense = cur
        ms.cell(row=cur, column=1, value="TOTAL EXPENSE").font = GRAND_FONT
        ms.cell(row=cur, column=1).fill = GRAND_FILL
        for col in range(2, total_col + 1):
            ltr  = get_column_letter(col)
            refs = "+".join(f"{ltr}{r}" for r in exp_subs)
            c = ms.cell(row=cur, column=col, value=f"={refs}")
            c.number_format = CURRENCY_FMT
            c.fill = GRAND_FILL
            c.font = GRAND_FONT
        ge_avg = ms.cell(row=cur, column=avg_col,
                         value=f"=AVERAGE({_avg_start}{cur}:{_avg_end}{cur})")
        ge_avg.number_format = CURRENCY_FMT
        ge_avg.fill = GRAND_FILL
        ge_avg.font = GRAND_FONT
        cur += 1

    # Net Income (Total Income − Total Expense) — per-month column formulas
    if ms_income_subtotal and ms_grand_expense:
        ms.cell(row=cur, column=1, value="NET INCOME").font = NET_FONT
        ms.cell(row=cur, column=1).fill = NET_FILL
        for col in range(2, total_col + 1):
            ltr = get_column_letter(col)
            c = ms.cell(row=cur, column=col,
                        value=f"={ltr}{ms_income_subtotal}-{ltr}{ms_grand_expense}")
            c.number_format = CURRENCY_FMT
            c.fill = NET_FILL
            c.font = NET_FONT
        ni_avg = ms.cell(row=cur, column=avg_col,
                         value=f"=AVERAGE({_avg_start}{cur}:{_avg_end}{cur})")
        ni_avg.number_format = CURRENCY_FMT
        ni_avg.fill = NET_FILL
        ni_avg.font = NET_FONT
        cur += 1

    # Stacked bar chart (expense categories only)
    all_exp_rows = ms_rental_rows + ms_personal_rows
    if all_exp_rows:
        try:
            ms._charts.clear()
            chart = BarChart()
            chart.type     = "col"
            chart.grouping = "stacked"
            _ym_lo, _ym_hi = active_months[0], active_months[-1]
            chart.title    = (
                f"Monthly Spending by Category  "
                f"{MONTHS[_ym_lo[1] - 1]} {_ym_lo[0]}–{MONTHS[_ym_hi[1] - 1]} {_ym_hi[0]}"
            )
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Month"
            chart.width    = 32
            chart.height   = 18
            _n_months = len(active_months)
            from openpyxl.chart import Series
            for r in all_exp_rows:
                series = Series(
                    Reference(ms, min_col=2, max_col=_n_months + 1, min_row=r, max_row=r),
                    title=ms.cell(row=r, column=1).value,
                )
                chart.series.append(series)
            chart.set_categories(Reference(ms, min_col=2, max_col=_n_months + 1, min_row=2))
            ms.add_chart(chart, f"A{cur + 1}")
        except Exception:
            pass

    # ── Per-month highlight: amber cell for largest category by magnitude ────
    # Legend in row 1 (to the right of "Category" header — only other content in that row)
    legend_cell = ms.cell(row=1, column=2,
                          value="🟡 = largest single category (by magnitude) that month")
    legend_cell.font = Font(name="Arial", italic=True, size=9, color="888888")

    if ms_cat_row_map:
        from collections import defaultdict
        monthly_totals: dict = defaultdict(lambda: defaultdict(float))
        for tx_row in tx_ws.iter_rows(min_row=2, values_only=True):
            date_val, _, _, cat, _, amount, include_in_net = (
                tx_row[0], tx_row[1], tx_row[2], tx_row[3], tx_row[4], tx_row[5], tx_row[6]
            )
            if not date_val or not cat or amount is None or include_in_net is False:
                continue
            if cat not in ms_cat_row_map:
                continue
            try:
                if hasattr(date_val, "year"):
                    row_date = date_val.date() if hasattr(date_val, "date") else date_val
                else:
                    row_date = date.fromisoformat(str(date_val))
            except Exception:
                continue
            monthly_totals[(row_date.year, row_date.month)][cat] += abs(float(amount))

        for col_idx, ym in enumerate(active_months, start=2):
            month_data = monthly_totals.get(ym, {})
            if not month_data:
                continue
            max_val = max(month_data.values())
            if max_val == 0:
                continue
            for cat, val in month_data.items():
                if val == max_val:
                    cell = ms.cell(row=ms_cat_row_map[cat], column=col_idx)
                    cell.fill = HIGHLIGHT_FILL
                    cell.font = HIGHLIGHT_FONT

    # ════════════════════════════════════════════════════════════════════════
    # YTD SUMMARY
    # ════════════════════════════════════════════════════════════════════════
    ytd = wb["YTD Summary"]
    _clear_ws(ytd, from_row=2)
    cur = 2

    ytd_income_subtotal   = None
    ytd_rental_subtotal   = None
    ytd_personal_subtotal = None
    ytd_grand_expense     = None

    def _ytd_sec_header(label):
        nonlocal cur
        ytd.cell(row=cur, column=1, value=label).font = SEC_FONT
        ytd.cell(row=cur, column=1).fill = SEC_FILL
        for col in [2, 3]:
            ytd.cell(row=cur, column=col).fill = SEC_FILL
        cur += 1

    def _ytd_cat_row(cat, tx_type):
        nonlocal cur
        r = cur
        ytd.cell(row=r, column=1, value=cat).font = Font(name="Arial", size=10)
        ytd.cell(row=r, column=2,
                 value=_sp(cat, tx_type, ytd_start, today)).number_format = CURRENCY_FMT
        cur += 1
        return r

    def _ytd_subtotal(label, data_rows, fill=SUB_FILL, font=SUB_FONT):
        nonlocal cur
        r = cur
        ytd.cell(row=r, column=1, value=label).font = font
        ytd.cell(row=r, column=1).fill = fill
        refs = "+".join(f"B{dr}" for dr in data_rows)
        c = ytd.cell(row=r, column=2, value=f"={refs}")
        c.number_format = CURRENCY_FMT
        c.fill = fill
        c.font = font
        cur += 1
        return r

    def _ytd_fill_pct(data_rows, denom_row, fill=None, font=None):
        for r in data_rows:
            c = ytd.cell(row=r, column=3,
                         value=f"=IF(B{denom_row}=0,\"\",B{r}/B{denom_row})")
            c.number_format = "0.0%"
            if fill: c.fill = fill
            if font: c.font = font

    # Income
    ytd_income_rows = []
    if income_cats:
        _ytd_sec_header("── INCOME ──")
        for cat in income_cats:
            ytd_income_rows.append(_ytd_cat_row(cat, "Income"))
        ytd_income_subtotal = _ytd_subtotal("TOTAL INCOME", ytd_income_rows)
        _ytd_fill_pct(ytd_income_rows, ytd_income_subtotal)
        ytd.cell(row=ytd_income_subtotal, column=3,
                 value="100%").number_format = "0.0%"
        ytd.cell(row=ytd_income_subtotal, column=3).font = SUB_FONT
        ytd.cell(row=ytd_income_subtotal, column=3).fill = SUB_FILL
        cur += 1

    # Rental Expense
    ytd_rental_rows = []
    if rental_exp_cats:
        _ytd_sec_header("── RENTAL EXPENSES ──")
        for cat in rental_exp_cats:
            ytd_rental_rows.append(_ytd_cat_row(cat, "Expense"))
        ytd_rental_subtotal = _ytd_subtotal("TOTAL RENTAL EXPENSE", ytd_rental_rows)
        cur += 1

    # Personal Expense
    ytd_personal_rows = []
    if personal_exp_cats:
        _ytd_sec_header("── PERSONAL EXPENSES ──")
        for cat in personal_exp_cats:
            ytd_personal_rows.append(_ytd_cat_row(cat, "Expense"))
        ytd_personal_subtotal = _ytd_subtotal("TOTAL PERSONAL EXPENSE", ytd_personal_rows)
        cur += 1

    # Grand Total Expense
    exp_subs_ytd = [r for r in [ytd_rental_subtotal, ytd_personal_subtotal] if r]
    if exp_subs_ytd:
        ytd_grand_expense = cur
        ytd.cell(row=cur, column=1, value="TOTAL EXPENSE").font = GRAND_FONT
        ytd.cell(row=cur, column=1).fill = GRAND_FILL
        refs = "+".join(f"B{r}" for r in exp_subs_ytd)
        c = ytd.cell(row=cur, column=2, value=f"={refs}")
        c.number_format = CURRENCY_FMT
        c.fill = GRAND_FILL
        c.font = GRAND_FONT
        cur += 1
        # Fill % for all expense category rows and their subtotals
        all_exp_ytd = ytd_rental_rows + ytd_personal_rows
        _ytd_fill_pct(all_exp_ytd, ytd_grand_expense)
        for r in exp_subs_ytd:
            _ytd_fill_pct([r], ytd_grand_expense, fill=SUB_FILL, font=SUB_FONT)

    # ── Net Summary block ─────────────────────────────────────────────────
    if ytd_income_subtotal and ytd_grand_expense:
        cur += 1  # blank separator
        net_items = [
            ("Total Income",          f"=B{ytd_income_subtotal}",                              CURRENCY_FMT),
            ("Total Expense",         f"=B{ytd_grand_expense}",                                CURRENCY_FMT),
            ("Net (Income − Expense)",f"=B{ytd_income_subtotal}-B{ytd_grand_expense}",         CURRENCY_FMT),
            ("Net as % of Income",    f"=IF(B{ytd_income_subtotal}=0,\"\","
                                      f"(B{ytd_income_subtotal}-B{ytd_grand_expense})"
                                      f"/B{ytd_income_subtotal})",                             "0.0%"),
        ]
        for label, formula, fmt in net_items:
            ytd.cell(row=cur, column=1, value=label).font = NET_FONT
            ytd.cell(row=cur, column=1).fill = NET_FILL
            c = ytd.cell(row=cur, column=2, value=formula)
            c.number_format = fmt
            c.fill = NET_FILL
            c.font = NET_FONT
            cur += 1

    # ════════════════════════════════════════════════════════════════════════
    # YoY TRENDS
    # ════════════════════════════════════════════════════════════════════════
    yoy = wb["YoY Trends"]
    _clear_ws(yoy, from_row=3)
    cur = 3

    all_cats_typed = (
        [(c, "Income")  for c in income_cats] +
        [(c, "Expense") for c in rental_exp_cats] +
        [(c, "Expense") for c in personal_exp_cats]
    )
    for cat, tx_type in all_cats_typed:
        yoy.cell(row=cur, column=1, value=cat).font = Font(name="Arial", size=10)
        yoy.cell(row=cur, column=2,
                 value=_sp(cat, tx_type, ytd_start, today)).number_format = CURRENCY_FMT
        yoy.cell(row=cur, column=3,
                 value=_sp(cat, tx_type, ly_start, ly_end)).number_format = CURRENCY_FMT
        yoy.cell(row=cur, column=4,
                 value=f"=B{cur}-C{cur}").number_format = CURRENCY_FMT
        yoy.cell(row=cur, column=5,
                 value=f"=IF(C{cur}=0,\"\",B{cur}/C{cur}-1)").number_format = "0.0%"
        cur += 1

    # ════════════════════════════════════════════════════════════════════════
    # DRILL-DOWN
    # ════════════════════════════════════════════════════════════════════════
    dd = wb["Drill-down"]
    _clear_ws(dd, from_row=1)
    dd.data_validations.dataValidation = []   # drop stale dropdowns before re-adding

    _helper_font = Font(name="Arial", size=9, italic=True, color="AAAAAA")
    _label_fill  = PatternFill("solid", fgColor="D6DCE4")
    _label_font  = Font(name="Arial", bold=True, size=10)
    _body_bold   = Font(name="Arial", bold=True, size=10)

    # ── Helper column I: months;  column J: categories ───────────────────
    # Written in gray italic so they recede visually.  DataValidation dropdowns
    # reference these ranges on the same sheet (most reliable in GS xlsx import).
    distinct_categories = income_cats + rental_exp_cats + personal_exp_cats
    month_vals = ["All Months"] + [f"M{yr}-{mn:02d}" for yr, mn in active_months]
    cat_vals   = ["All Categories"] + distinct_categories

    for i, val in enumerate(month_vals, start=1):
        dd.cell(row=i, column=14, value=val).font = _helper_font

    for j, val in enumerate(cat_vals, start=1):
        dd.cell(row=j, column=15, value=val).font = _helper_font

    month_dv_range = f"$N$1:$N${len(month_vals)}"
    cat_dv_range   = f"$O$1:$O${len(cat_vals)}"

    # ── Section 1: Quick Search ──────────────────────────────────────────
    # Row 1: section header band
    for col in range(1, 7):
        c = dd.cell(row=1, column=col)
        c.fill  = SEC_FILL
        c.font  = SEC_FONT if col == 1 else Font(name="Arial", size=10)
        c.value = "Quick Search" if col == 1 else None

    # Row 2: search term (B2), optional Year filter (D2), optional Month filter (F2).
    # B2, D2, F2 left blank — user types here.  Month (M-format) takes precedence
    # over Year when both are filled; blank = no time filter on that dimension.
    dd.cell(row=2, column=1, value="Search term:").font = _body_bold
    dd.cell(row=2, column=3, value="Year:").font         = _body_bold
    dd.cell(row=2, column=5, value="Month:").font        = _body_bold

    # Row 3: column headers above QUERY output
    for col_idx, label in enumerate(
        ["Date", "Description", "Account", "Category", "Type", "Amount"], start=1
    ):
        c = dd.cell(row=3, column=col_idx, value=label)
        c.fill = _label_fill
        c.font = _label_font

    # Row 4: QUERY formula — spills results when B2 has content, "" when empty.
    # D2 = optional year (e.g. "2026"); F2 = optional month (e.g. "M2026-05").
    # Month takes precedence over Year. Both blank = no Col1 condition at all.
    dd.cell(row=4, column=1,
            value='=IF(B2="","",IFERROR(QUERY(Transactions!A2:F10000,'
                  '"select * where Col1 is not null and lower(Col2) like \'%"'
                  '&LOWER(B2)&"%\'"'
                  '&IF(F2<>""," and Col1 starts with \'"&MID(F2,2,7)&"\'",IF(D2<>""," and Col1 starts with \'"&D2&"\'","")),'
                  '0),"No matching transactions"))')

    # ── Section 2: Month + Category Lookup ──────────────────────────────
    # Row 30: section header band — H:M (Section 2 occupies H30:M33+)
    for col in range(8, 14):
        c = dd.cell(row=30, column=col)
        c.fill  = SEC_FILL
        c.font  = SEC_FONT if col == 8 else Font(name="Arial", size=10)
        c.value = "Month & Category Lookup" if col == 8 else None

    # Row 31: dropdown input labels + cells (H31="Month:", I31=dropdown; K31="Category:", L31=dropdown)
    dd.cell(row=31, column=8,  value="Month:").font    = _body_bold   # H31
    dd.cell(row=31, column=11, value="Category:").font = _body_bold   # K31
    # I31 and L31 left blank — populated at runtime via DataValidation dropdowns

    # Row 32: column headers above QUERY output
    for col_idx, label in enumerate(
        ["Date", "Description", "Account", "Category", "Type", "Amount"], start=1
    ):
        c = dd.cell(row=32, column=col_idx + 7, value=label)
        c.fill = _label_fill
        c.font = _label_font

    # Row 33: QUERY formula — WHERE clause built dynamically from I31/L31.
    # Month format "M{year}-{month:02d}" (e.g. "M2026-06"); MID(I31,2,7) strips
    # the "M" prefix to get "2026-06" for GQL "starts with" text matching.
    # Dates in Transactions are stored as text strings ("2026-06-22"), so
    # year()/month() GQL functions cannot be used — they require a true Date column.
    dd.cell(row=33, column=8,
            value='=IFERROR(QUERY(Transactions!A2:F10000,'
                  '"select * where Col1 is not null"'
                  '&IF(I31="All Months",""," and Col1 starts with \'"&MID(I31,2,7)&"\'")'
                  '&IF(L31="All Categories",""," and Col4 = \'"&L31&"\'"),'
                  '0),"No matching transactions")')

    # ── DataValidation dropdowns ─────────────────────────────────────────
    # showDropDown=False → arrow IS visible (openpyxl name is inverted).
    # showErrorMessage=True + default errorStyle "stop" → rejects invalid entries.
    dv_month = DataValidation(
        type="list", formula1=month_dv_range, showDropDown=False,
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid month", error="Select a value from the dropdown list.",
    )
    dv_cat = DataValidation(
        type="list", formula1=cat_dv_range, showDropDown=False,
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid category", error="Select a value from the dropdown list.",
    )
    dd.add_data_validation(dv_month)
    dd.add_data_validation(dv_cat)
    dv_month.add("I31")
    dv_cat.add("L31")


def write_spending_ledger(filepath: str, new_transactions: list) -> dict:
    year = date.today().year

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        if "Transactions"    not in wb.sheetnames: wb.create_sheet("Transactions", 0)
        if "Monthly Summary" not in wb.sheetnames: _build_summary_sheet(wb, year)
        if "YTD Summary"     not in wb.sheetnames: _build_ytd_sheet(wb)
        if "YoY Trends"      not in wb.sheetnames: _build_yoy_sheet(wb)
        if "Drill-down"      not in wb.sheetnames: _build_drilldown_sheet(wb)
        # Migrate: add IncludeInNet column G to existing files that pre-date this column
        _tx = wb["Transactions"]
        if _tx["G1"].value != "IncludeInNet":
            _tx["G1"].value = "IncludeInNet"
            _tx["G1"].fill  = HEADER_FILL
            _tx["G1"].font  = HEADER_FONT
            _tx["G1"].alignment = Alignment(horizontal="center")
            _tx.column_dimensions["G"].width = 13
            for row in _tx.iter_rows(min_row=2):
                if row[0].value:  # only rows that have data
                    row[6].value = True
        # Migrate: add SourceRef column H for CSV-import dedup (existing rows left blank)
        if _tx["H1"].value != "SourceRef":
            _tx["H1"].value = "SourceRef"
            _tx["H1"].fill  = HEADER_FILL
            _tx["H1"].font  = HEADER_FONT
            _tx["H1"].alignment = Alignment(horizontal="center")
            _tx.column_dimensions["H"].width = 32
    else:
        wb = Workbook()
        wb.remove(wb.active)
        tx_ws = wb.create_sheet("Transactions")
        _header_row(tx_ws, ["Date", "Description", "Account", "Category", "Type", "Amount", "IncludeInNet", "SourceRef"])
        tx_ws.column_dimensions["A"].width = 12
        tx_ws.column_dimensions["B"].width = 38
        tx_ws.column_dimensions["C"].width = 18
        tx_ws.column_dimensions["D"].width = 34
        tx_ws.column_dimensions["E"].width = 10
        tx_ws.column_dimensions["F"].width = 14
        tx_ws.column_dimensions["G"].width = 13
        tx_ws.column_dimensions["H"].width = 32
        tx_ws.freeze_panes = "A2"
        _build_summary_sheet(wb, year)
        _build_ytd_sheet(wb)
        _build_yoy_sheet(wb)
        _build_drilldown_sheet(wb)

    tx_ws = wb["Transactions"]
    existing_keys = _existing_keys(tx_ws)

    added = skipped = 0
    for tx in sorted(new_transactions, key=lambda t: t.get("date", ""), reverse=True):
        key = _tx_dedup_key(tx)
        if key in existing_keys:
            skipped += 1
            continue

        cat            = tx.get("category", "Uncategorized")
        include_in_net = not tx.get("exclude_from_net", False)
        row_idx  = tx_ws.max_row + 1
        row_fill = _category_fill(cat)
        for col_idx, val in enumerate(
            [tx.get("date"), tx.get("name", ""), tx.get("account_label", ""),
             cat, tx.get("type", "Expense"), tx.get("amount", 0.0), include_in_net,
             tx.get("transaction_id", "")], 1
        ):
            c = tx_ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = BODY_FONT
            c.fill = row_fill
            if col_idx == 6:
                c.number_format = CURRENCY_FMT

        existing_keys.add(key)
        added += 1

    _refresh_summary_formulas(wb, year)

    wb.save(filepath)
    return {"added": added, "skipped": skipped}


def get_last_snapshot_month(wb) -> str | None:
    """Read the last year-month a snapshot email was sent from the hidden _Meta sheet.
    Returns None if the sheet or key is absent (triggers snapshot on next cloud sync).
    Compatible with workbooks opened read_only=True.
    """
    if "_Meta" not in wb.sheetnames:
        return None
    ws = wb["_Meta"]
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if row and row[0] == "last_snapshot_month":
            return str(row[1]) if row[1] else None
    return None


def set_last_snapshot_month(wb, year_month: str) -> None:
    """Write the year-month marker into the hidden _Meta sheet, creating it if absent.
    Caller must save the workbook after this call.
    """
    if "_Meta" not in wb.sheetnames:
        ws = wb.create_sheet("_Meta")
        ws.sheet_state = "hidden"
    else:
        ws = wb["_Meta"]
    ws["A1"] = "last_snapshot_month"
    ws["B1"] = year_month
