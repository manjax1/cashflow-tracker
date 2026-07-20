#!/usr/bin/env python3
"""Rename a category across the ledger (Transactions + Splits sheets).
Deterministic, no API — for fixing typo categories or merges.

Usage:
    python scripts/rename_category.py "Grociers" "Groceries"
    python scripts/rename_category.py "Grociers" "Groceries" --keyword GFLP   # only rows whose description contains GFLP
"""

import os
import shutil
import sys
from datetime import datetime

import openpyxl
from dotenv import load_dotenv

load_dotenv()

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LEDGER = os.environ.get("SPENDING_LEDGER_FILE_PATH", os.path.join(ROOT, "cashflow-tracker.xlsx"))


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    old, new = sys.argv[1], sys.argv[2]
    kw = None
    if "--keyword" in sys.argv:
        kw = sys.argv[sys.argv.index("--keyword") + 1].lower()

    backup = LEDGER.replace(".xlsx", f"_BACKUP_{datetime.now():%Y-%m-%d_%H%M%S}.xlsx")
    shutil.copy2(LEDGER, backup)
    wb = openpyxl.load_workbook(LEDGER)
    changed = 0

    ws = wb["Transactions"]
    header = [c.value for c in ws[1]]
    cat_i = header.index("Category")
    desc_i = header.index("Description")
    for row in ws.iter_rows(min_row=2):
        if row[cat_i].value == old and (kw is None or kw in str(row[desc_i].value).lower()):
            row[cat_i].value = new
            changed += 1

    if "Splits" in wb.sheetnames:
        sh = wb["Splits"]
        shh = [c.value for c in sh[1]]
        sc = shh.index("Category")
        for row in sh.iter_rows(min_row=2):
            if row[sc].value == old:
                row[sc].value = new
                changed += 1

    wb.save(LEDGER)
    print(f"Renamed '{old}' -> '{new}' in {changed} rows"
          + (f" (keyword filter: {kw})" if kw else "")
          + f". Backup: {os.path.basename(backup)}")
    print("Next: 'push' from the agent CLI to sync Drive.")


if __name__ == "__main__":
    main()
